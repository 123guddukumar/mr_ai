import os
import instaloader
import openpyxl

# Initialize Instaloader instance
L = instaloader.Instaloader(
    download_pictures=False,          # Skip downloading images
    download_videos=True,             # Download videos
    download_video_thumbnails=False,  # Skip video thumbnails
    download_geotags=False,           # Skip location data
    download_comments=False,          # Skip comments
    save_metadata=False,              # Skip JSON metadata files
    post_metadata_txt_pattern=""      # Skip caption text files
)

# Optional: Log in if you encounter private links or rate limits
# L.login("your_username", "your_password")

def download_from_custom_excel(filename="links.xlsx"):
    """
    Is function se Column A me maujood saare Instagram links automatic extract ho jayenge.
    """
    if not os.path.exists(filename):
        print(f"Error: '{filename}' nahi mila. Kripya sahi file name check karein.")
        return

    links = []
    try:
        # Load workbook and select active sheet using openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active

        # Column A (1st column) ki har row ko check karein
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            
            if cell_value and isinstance(cell_value, str):
                cell_value = cell_value.strip()
                # Check agar cell me instagram ka valid link hai
                if "instagram.com" in cell_value.lower():
                    links.append(cell_value)

    except Exception as e:
        print(f"Excel file read karne me dikkat aayi: {e}")
        return

    print(f"Excel sheet se kul {len(links)} Instagram links mile.\n")

    for url in links:
        try:
            # Extract the shortcode from the URL
            parts = url.rstrip("/").split("/")
            
            # Query parameters (?igsh=...) ko hatane ke liye split karein
            main_url_part = parts[-1].split("?")[0]
            shortcode = main_url_part
            
            # Agar URL ke last me reel/p/reels ho to usse pehle wala part shortcode hoga
            if shortcode in ["reels", "p", "reel", ""]:
                shortcode = parts[-2].split("?")[0]
            
            print(f"Processing shortcode: {shortcode}")
            
            # Fetch the post and download it
            post = instaloader.Post.from_shortcode(L.context, shortcode)
            
            if post.is_video:
                # Video download target folder
                L.download_post(post, target=f"downloaded_{shortcode}")
                print(f"Successfully downloaded: {shortcode}")
            else:
                print(f"Skipped: Post {shortcode} video nahi hai.")
                
        except Exception as e:
            print(f"Error processing {url}: {e}")

if __name__ == "__main__":
    # Apni file ka sahi naam yahan likhein (e.g., 'links.xlsx')
    download_from_custom_excel(filename="links.xlsx")
