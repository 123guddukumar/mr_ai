"""
MR AI RAG - Classroom & Exam Management Routes
"""

import secrets
import logging
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Header, BackgroundTasks, Form, UploadFile, File
import asyncio
import io
import openpyxl
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, selectinload

from app.core.database import get_db
from app.core.clients import validate_client_token
from app.core.models import Exam, PaperClassroom, Subject, ChapterClassroom, TopicClassroom, SubtopicClassroom, CurrentAffairTopic, CurrentAffairReel, PYQSet, PYQQuestion, PaperChat, PYQChat
import os

logger = logging.getLogger(__name__)
router = APIRouter()


def extract_text_from_pdf(contents: bytes, max_pages: int = 30) -> str:
    """
    Extract text from PDF using PyMuPDF (fitz), pdfplumber, or PyPDF2 as fallback.
    Returns the extracted text, or empty string if all failed/no text found.
    """
    # Try 1: PyMuPDF (fitz) - installed in virtual env & very robust
    try:
        import fitz
        doc = fitz.open(stream=contents, filetype="pdf")
        pages = []
        for i in range(min(len(doc), max_pages)):
            page = doc.load_page(i)
            text = page.get_text()
            if text and text.strip():
                pages.append(text)
        txt = "\n".join(pages).strip()
        if txt:
            logger.info(f"Successfully extracted {len(txt)} chars from PDF using PyMuPDF (fitz).")
            return txt
    except Exception as e:
        logger.warning(f"PyMuPDF PDF extraction failed: {e}")

    # Try 2: pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            pages = []
            for page in pdf.pages[:max_pages]:
                pt = page.extract_text()
                if pt and pt.strip():
                    pages.append(pt)
            txt = "\n".join(pages).strip()
            if txt:
                logger.info(f"Successfully extracted {len(txt)} chars from PDF using pdfplumber.")
                return txt
    except Exception as e:
        logger.warning(f"pdfplumber PDF extraction failed: {e}")

    # Try 3: PyPDF2
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(contents))
        pages = []
        for page_num in range(min(len(reader.pages), max_pages)):
            page = reader.pages[page_num]
            pt = page.extract_text()
            if pt and pt.strip():
                pages.append(pt)
        txt = "\n".join(pages).strip()
        if txt:
            logger.info(f"Successfully extracted {len(txt)} chars from PDF using PyPDF2.")
            return txt
    except Exception as e:
        logger.warning(f"PyPDF2 PDF extraction failed: {e}")

    return ""


async def extract_text_from_scanned_pdf(contents: bytes, max_pages: int = 15) -> str:
    """
    Renders PDF pages as images and sends them to OpenAI (gpt-4o-mini) to extract the full text.
    """
    import base64
    import fitz  # PyMuPDF
    from openai import AsyncOpenAI
    from app.core.config import settings

    api_key = settings.OPENAI_API_KEY
    if not api_key:
        logger.warning("OpenAI API key missing, cannot perform multimodal OCR text extraction.")
        return ""

    try:
        doc = fitz.open(stream=contents, filetype="pdf")
    except Exception as e:
        logger.error(f"Failed to open PDF for OCR rendering: {e}")
        return ""

    content = [
        {
            "type": "text",
            "text": "Extract all readable text from the uploaded exam/document pages. Perform clean OCR. Do not summarize or explain, just output the exact extracted text from the pages."
        }
    ]

    rendered_count = 0
    for i in range(min(len(doc), max_pages)):
        try:
            page = doc.load_page(i)
            # Use 120 DPI to optimize payload size
            pix = page.get_pixmap(dpi=120)
            png_bytes = pix.tobytes("png")
            b64_data = base64.b64encode(png_bytes).decode("utf-8")
            content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{b64_data}"
                }
            })
            rendered_count += 1
        except Exception as err:
            logger.error(f"Failed to render page {i} for OCR: {err}")

    if rendered_count == 0:
        return ""

    logger.info(f"Sending {rendered_count} pages to OpenAI (gpt-4o-mini) for scanned PDF text OCR...")
    
    try:
        client = AsyncOpenAI(api_key=api_key)
        resp = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": content}],
            max_tokens=2048,
            temperature=0.1
        )
        text = resp.choices[0].message.content.strip()
        logger.info(f"Successfully OCR'd {len(text)} characters from scanned PDF using OpenAI.")
        return text
    except Exception as e:
        logger.error(f"OpenAI text OCR failed: {e}")
    return ""


async def extract_mcqs_from_scanned_pdf(contents: bytes, max_pages: int = 15) -> list:
    """
    Renders PDF pages as images and sends them to Groq (llama-3.2-11b-vision-preview) or OpenAI (gpt-4o-mini)
    for multimodal MCQ extraction.
    """
    import base64
    import fitz  # PyMuPDF
    import json
    from openai import AsyncOpenAI
    from app.core.config import settings

    try:
        doc = fitz.open(stream=contents, filetype="pdf")
    except Exception as e:
        logger.error(f"Failed to open PDF for MCQ OCR rendering: {e}")
        return []

    prompt = (
        "Extract all Multiple Choice Questions (MCQs) from the uploaded exam pages.\n"
        "Return a JSON array containing the questions. "
        "Each question object MUST have the following structure:\n"
        "{\n"
        "  \"question\": \"Question text\",\n"
        "  \"options\": [\"Option A\", \"Option B\", \"Option C\", \"Option D\"],\n"
        "  \"correct\": \"Correct option text or letter\"\n"
        "}\n"
        "If the correct answer is not explicitly marked on the page, leave \"correct\" as empty string.\n"
        "Return a JSON list directly."
    )

    content = [{"type": "text", "text": prompt}]

    rendered_count = 0
    for i in range(min(len(doc), max_pages)):
        try:
            page = doc.load_page(i)
            # Render page to a pixmap (DPI 120 is a good balance of quality and size for Groq/OpenAI vision)
            pix = page.get_pixmap(dpi=120)
            png_bytes = pix.tobytes("png")
            b64_data = base64.b64encode(png_bytes).decode("utf-8")
            content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{b64_data}"
                }
            })
            rendered_count += 1
        except Exception as err:
            logger.error(f"Failed to render page {i} for MCQ OCR: {err}")

    if rendered_count == 0:
        logger.warning("No pages could be rendered as images for MCQ OCR.")
        return []

    # Try 1: Groq Vision (llama-3.2-11b-vision-preview) as requested
    groq_api_key = settings.GROQ_API_KEY
    if groq_api_key:
        logger.info(f"Sending {rendered_count} pages to Groq (llama-3.2-11b-vision-preview) for MCQ OCR...")
        try:
            client = AsyncOpenAI(api_key=groq_api_key, base_url="https://api.groq.com/openai/v1")
            resp = await client.chat.completions.create(
                model="llama-3.2-11b-vision-preview",
                messages=[{"role": "user", "content": content}],
                max_tokens=4096,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            text_response = resp.choices[0].message.content.strip()
            parsed = json.loads(text_response)
            
            if isinstance(parsed, dict):
                for val in parsed.values():
                    if isinstance(val, list):
                        logger.info(f"Successfully extracted {len(val)} questions from scanned PDF via Groq vision OCR.")
                        return val
            elif isinstance(parsed, list):
                logger.info(f"Successfully extracted {len(parsed)} questions from scanned PDF via Groq vision OCR.")
                return parsed
        except Exception as groq_ocr_err:
            logger.warning(f"Groq Vision MCQ OCR failed: {groq_ocr_err}. Falling back to OpenAI...")

    # Try 2: OpenAI (gpt-4o-mini) as fallback
    openai_api_key = settings.OPENAI_API_KEY
    if openai_api_key:
        logger.info(f"Sending {rendered_count} pages to OpenAI (gpt-4o-mini) for MCQ OCR fallback...")
        try:
            client = AsyncOpenAI(api_key=openai_api_key)
            resp = await client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": content}],
                max_tokens=4096,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            text_response = resp.choices[0].message.content.strip()
            parsed = json.loads(text_response)
            
            if isinstance(parsed, dict):
                for val in parsed.values():
                    if isinstance(val, list):
                        logger.info(f"Successfully extracted {len(val)} questions from scanned PDF via OpenAI OCR.")
                        return val
            elif isinstance(parsed, list):
                logger.info(f"Successfully extracted {len(parsed)} questions from scanned PDF via OpenAI OCR.")
                return parsed
        except Exception as oai_ocr_err:
            logger.error(f"OpenAI MCQ OCR fallback failed: {oai_ocr_err}")

    return []


async def generate_groq_response(prompt: str, system_prompt: str = "You are a helpful assistant.") -> str:
    """Explicitly generates a response using Groq (llama-3.3-70b-versatile)."""
    from app.core.config import settings
    from openai import AsyncOpenAI
    api_key = settings.GROQ_API_KEY
    model = settings.GROQ_MODEL or "llama-3.3-70b-versatile"
    if not api_key:
        raise RuntimeError("Groq API key is not configured in settings.")
    client = AsyncOpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")
    resp = await client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1024,
        temperature=0.3
    )
    return resp.choices[0].message.content.strip()


# ── Auth dependency ───────────────────────────────────────────────────────────

def _require_client(
    x_app_token: Optional[str] = Header(None, alias="X-App-Token"),
    token: Optional[str] = None,
    db: Session = Depends(get_db),
) -> dict:
    actual_token = x_app_token or token
    if not actual_token:
        raise HTTPException(401, "Missing X-App-Token header or token parameter.")
    record = validate_client_token(actual_token, db=db)
    if not record:
        raise HTTPException(401, "Invalid or expired token. Please login again.")
    return record



# ── Pydantic Request Models ───────────────────────────────────────────────────

class CreateExamReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    image_url: Optional[str] = ""
    description: Optional[str] = ""
    category: Optional[str] = ""

class CreatePaperReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)

class CreateSubjectReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    color: Optional[str] = "#4f46e5"
    image_url: Optional[str] = ""

class CreateChapterReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    image_url: Optional[str] = ""

class GenerateImageReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=500)
    type: str  # "subject" or "chapter"
    context: Optional[str] = ""

class RegenerateImageReq(BaseModel):
    entity_id: str
    type: str  # "subject" | "chapter" | "topic" | "subtopic" | "subtopic_banner" | "topic_banner"

class GeneratePromptReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=500)
    type: str  # "subject" | "chapter" | "topic" | "subtopic" | "subtopic_banner"
    context: Optional[str] = ""

def fetch_wikimedia_image_url(keyword: str) -> Optional[str]:
    import httpx
    import urllib.parse
    headers = {"User-Agent": "ClassroomImageGen/1.0 (admin@mr-ai.com)"}
    try:
        cleaned = urllib.parse.quote(keyword.replace("/", " "))
        url = f"https://commons.wikimedia.org/w/api.php?action=query&generator=search&gsrnamespace=6&gsrsearch={cleaned}&gsrlimit=10&prop=imageinfo&iiprop=url&format=json"
        with httpx.Client(timeout=10.0) as client:
            resp = client.get(url, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                pages = data.get("query", {}).get("pages", {})
                for page_id, page_info in pages.items():
                    image_info = page_info.get("imageinfo", [])
                    if image_info:
                        img_url = image_info[0].get("url", "")
                        lower_url = img_url.lower()
                        if lower_url.endswith((".jpg", ".jpeg", ".png")):
                            return img_url
    except Exception as e:
        logger.warning(f"Error searching Wikimedia for keyword '{keyword}': {e}")
    return None

async def generate_educational_image_prompt(title: str, subtitle: str = "Subject", context: str = "") -> str:
    """
    Generates a highly detailed, professional educational image description.
    Returns ONLY visual description — no aspect ratio, no format, no size instructions.
    The calling function (extension or Pollinations URL) handles format/size separately.
    """
    from app.services.llm import generate_simple_response
    
    system_prompt = (
        "You are an expert visual artist and prompt engineer for AI image generators (Midjourney, Flux, SDXL).\n"
        "Write a vivid, detailed image description for an educational illustration.\n"
        "Rules:\n"
        "- Output ONLY the visual description. No intro, no outro, no explanations.\n"
        "- Do NOT mention aspect ratio, image size, format, width, height, 9:16, 1:1, 16:9, portrait, landscape.\n"
        "- Do NOT include text, words, titles, letters in the described image.\n"
        "- Focus on objects, colors, lighting, textures, composition.\n"
        "- Be specific and vivid — name exact objects, monuments, scientific instruments, etc."
    )
    
    is_banner = "banner" in subtitle.lower()
    context_str = f" for {context} Exam Preparation" if context else ""
    
    if is_banner:
        style_hint = "wide cinematic panoramic composition, ultra-wide landscape view"
    else:
        style_hint = "centered square composition, symmetrical layout, hero element in center"
    
    user_prompt = (
        f"Write a detailed image generation description for:\n"
        f"Subject: {title}\n"
        f"Context: {context}\n\n"
        f"Style: {style_hint}\n\n"
        f"The description must:\n"
        f"- Feature 4-6 specific symbolic objects/structures representing '{title}' (e.g., for History: ancient stone pillars, Mughal dome, spinning wheel, ruined fort arches; for Geography: Himalayan peaks, globe, compass, river map; for Science: DNA helix, telescope, test tubes, atomic model)\n"
        f"- Specify 2-3 harmonious colors (e.g., saffron gold and navy blue, or forest green and terracotta)\n"
        f"- Include a specific texture/material (parchment, marble, dark stone, leather atlas)\n"
        f"- Describe dramatic cinematic lighting (god rays, golden hour, soft studio)\n"
        f"- End with: ultra-detailed 4K photorealistic artwork, premium educational illustration, no text, no words\n\n"
        f"Output the description directly. No brackets, no instructions, no format words."
    )
    
    try:
        prompt = await generate_simple_response(user_prompt, system_prompt=system_prompt, max_tokens=300)
        prompt = prompt.strip().strip('"').strip("'")
        if prompt and len(prompt) > 30:
            logger.info(f"Generated educational image prompt: {prompt[:100]}...")
            return prompt
    except Exception as e:
        logger.warning(f"LLM prompt generation failed, using fallback: {e}")
    
    # Fallback — pure visual description, no format/size words
    if is_banner:
        return (
            f"Panoramic cinematic educational illustration representing {title}{context_str}. "
            f"Wide landscape scene with symbolic elements of {title}, dramatic golden hour lighting, "
            f"rich colors, ultra-detailed 4K photorealistic artwork, no text, no words"
        )
    else:
        return (
            f"Centered photorealistic educational illustration representing {title}{context_str}. "
            f"Grand collage of symbolic objects related to {title}, dramatic cinematic lighting, "
            f"rich colors, premium artwork, ultra-detailed 4K quality, no text, no words"
        )


async def generate_ai_image_and_upload(title: str, subtitle: str = "Subject", context: str = "") -> str:
    import urllib.parse
    import secrets
    import httpx
    import os
    import time
    
    # Construct a professional realistic prompt based on the title & context using LLM
    prompt = await generate_educational_image_prompt(title, subtitle=subtitle, context=context)
    encoded = urllib.parse.quote(prompt)
    seed = secrets.token_hex(4)
    # Use Flux or standard pollinations
    image_url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&seed={seed}&model=flux"
    
    filename = f"classroom_{secrets.token_hex(8)}.jpg"
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads", "images")
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, filename)
    
    max_retries = 20
    retry_delay = 3.0 # seconds
    
    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"Attempting to fetch image from Pollinations AI (Attempt {attempt}/{max_retries})...")
            with httpx.Client(timeout=45.0) as client:
                resp = client.get(image_url)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    with open(filepath, "wb") as f:
                        f.write(resp.content)
                    
                    # Upload to R2 if configured
                    try:
                        from app.services.r2_storage import upload_to_r2
                        r2_key = f"classroom/images/{filename}"
                        r2_url = upload_to_r2(filepath, r2_key, "image/jpeg")
                        if r2_url:
                            return r2_url
                    except Exception as r2_err:
                        logger.error(f"R2 upload failed for AI generated image: {r2_err}")
                    
                    return f"/uploads/images/{filename}"
                else:
                    logger.warning(f"Pollinations AI returned status {resp.status_code} (attempt {attempt}/{max_retries})")
        except Exception as e:
            logger.warning(f"Error fetching from Pollinations AI on attempt {attempt}: {e}")
        
        # If not successful, wait and retry
        if attempt < max_retries:
            logger.info(f"Waiting {retry_delay}s before retrying Pollinations AI...")
            time.sleep(retry_delay)
        
    # If AI generation fails, fall back to locally drawn premium image
    return generate_premium_image_locally(title, subtitle)


async def generate_banner_image_and_upload(title: str, subtitle: str = "Subtopic Banner", context: str = "") -> str:
    import urllib.parse
    import secrets
    import httpx
    import os
    import time
    
    # Construct a professional realistic prompt based on the title & context using LLM
    prompt = await generate_educational_image_prompt(title, subtitle=subtitle, context=context)
    encoded = urllib.parse.quote(prompt)
    seed = secrets.token_hex(4)
    # Use Flux, 16:9 ratio (2560x1440 high resolution)
    image_url = f"https://image.pollinations.ai/prompt/{encoded}?width=2560&height=1440&nologo=true&seed={seed}&model=flux"
    
    filename = f"banner_{secrets.token_hex(8)}.jpg"
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads", "images")
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, filename)
    
    max_retries = 20
    retry_delay = 3.0 # seconds
    
    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"Attempting to fetch banner from Pollinations AI (Attempt {attempt}/{max_retries})...")
            with httpx.Client(timeout=45.0) as client:
                resp = client.get(image_url)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    with open(filepath, "wb") as f:
                        f.write(resp.content)
                    
                    # Upload to R2 if configured
                    try:
                        from app.services.r2_storage import upload_to_r2
                        r2_key = f"classroom/banners/{filename}"
                        r2_url = upload_to_r2(filepath, r2_key, "image/jpeg")
                        if r2_url:
                            return r2_url
                    except Exception as r2_err:
                        logger.error(f"R2 upload failed for AI generated banner: {r2_err}")
                    
                    return f"/uploads/images/{filename}"
                else:
                    logger.warning(f"Pollinations AI returned status {resp.status_code} for banner (attempt {attempt}/{max_retries})")
        except Exception as e:
            logger.warning(f"Error fetching banner from Pollinations AI on attempt {attempt}: {e}")
            
        if attempt < max_retries:
            logger.info(f"Waiting {retry_delay}s before retrying Pollinations AI banner...")
            time.sleep(retry_delay)
            
    return generate_premium_banner_locally(title, subtitle)


def generate_premium_banner_locally(title: str, subtitle: str = "Subtopic Banner") -> str:
    from PIL import Image, ImageDraw, ImageFont, ImageOps
    import os
    import secrets
    import httpx
    
    w, h = 1280, 720
    img = None
    headers = {"User-Agent": "ClassroomImageGen/1.0 (admin@mr-ai.com)"}
    
    # 1. Search Wikimedia Commons for a relevant illustration matching the title
    wikimedia_url = fetch_wikimedia_image_url(title)
    if wikimedia_url:
        try:
            with httpx.Client(timeout=15.0) as client:
                resp = client.get(wikimedia_url, headers=headers)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    from io import BytesIO
                    raw_img = Image.open(BytesIO(resp.content))
                    # Crop and fit image centered
                    img = ImageOps.fit(raw_img, (w, h)).convert("RGB")
        except Exception as e:
            logger.warning(f"Failed to load Wikimedia image '{wikimedia_url}' for banner: {e}")
            
    # 2. Fallback to royal linear gradient background if download failed or no match found
    if img is None:
        img = Image.new("RGB", (w, h), "#0b192c")
        draw = ImageDraw.Draw(img)
        for y in range(h):
            r = int(11 + (30 - 11) * (y / h))
            g = int(25 + (62 - 25) * (y / h))
            b = int(44 + (98 - 44) * (y / h))
            draw.line([(0, y), (w, y)], fill=(r, g, b))
    else:
        # Create a beautiful dark blue transparent glassmorphic overlay for contrast
        overlay = Image.new("RGBA", (w, h), (11, 25, 44, 185)) # Navy overlay (approx 72% opacity)
        img.paste(overlay, (0, 0), overlay)
        
    draw = ImageDraw.Draw(img)
    border_color = "#D4AF37"  # Gold
    
    # Outer elegant frame
    draw.rectangle([(30, 30), (w - 30, h - 30)], outline=border_color, width=4)
    draw.rectangle([(38, 38), (w - 38, h - 38)], outline=border_color, width=1)
    
    # Corner ornaments
    draw.line([(20, 50), (60, 50)], fill=border_color, width=3)
    draw.line([(50, 20), (50, 60)], fill=border_color, width=3)
    
    draw.line([(w - 60, 50), (w - 20, 50)], fill=border_color, width=3)
    draw.line([(w - 50, 20), (w - 50, 60)], fill=border_color, width=3)
    
    draw.line([(20, h - 50), (60, h - 50)], fill=border_color, width=3)
    draw.line([(50, h - 60), (50, h - 20)], fill=border_color, width=3)
    
    draw.line([(w - 60, h - 50), (w - 20, h - 50)], fill=border_color, width=3)
    draw.line([(w - 50, h - 60), (w - 50, h - 20)], fill=border_color, width=3)
    
    # Font path
    font_path = "C:\\Windows\\Fonts\\georgiab.ttf"
    if not os.path.exists(font_path):
        font_path = "C:\\Windows\\Fonts\\timesbd.ttf"
    if not os.path.exists(font_path):
        font_path = "C:\\Windows\\Fonts\\arial.ttf"
        
    try:
        title_font = ImageFont.truetype(font_path, 54)
        sub_font = ImageFont.truetype(font_path, 24)
    except Exception:
        title_font = ImageFont.load_default()
        sub_font = ImageFont.load_default()
        
    # Central geometric medallion outline
    draw.ellipse([(w//2 - 140, h//2 - 160), (w//2 + 140, h//2 + 80)], outline=border_color, width=2)
    
    # Center coordinates of the medallion
    cx, cy = w // 2, h // 2 - 40
    title_lower = title.lower()
    
    # Draw keyword-specific gold vector icons
    if any(k in title_lower for k in ["history", "itihas", "ancient", "medieval", "modern", "war", "struggle", "movement", "period", "empire"]):
        draw.line([(cx, cy - 35), (cx, cy + 35)], fill=border_color, width=3) # Spine
        draw.polygon([(cx, cy - 35), (cx - 45, cy - 28), (cx - 45, cy + 28), (cx, cy + 35)], outline=border_color, width=2) # Left
        draw.polygon([(cx, cy - 35), (cx + 45, cy - 28), (cx + 45, cy + 28), (cx, cy + 35)], outline=border_color, width=2) # Right
        
    elif any(k in title_lower for k in ["geography", "bhugol", "climate", "resource", "earth", "settlement", "map", "space", "astronomy", "physical", "atmosphere"]):
        r = 40
        draw.ellipse([(cx - r, cy - r), (cx + r, cy + r)], outline=border_color, width=2)
        draw.line([(cx - r, cy), (cx + r, cy)], fill=border_color, width=1)
        draw.line([(cx, cy - r), (cx, cy + r)], fill=border_color, width=1)
        draw.ellipse([(cx - r, cy - r//2), (cx + r, cy + r//2)], outline=border_color, width=1)
        draw.ellipse([(cx - r//2, cy - r), (cx + r//2, cy + r)], outline=border_color, width=1)
        
    elif any(k in title_lower for k in ["polity", "constitution", "samvidhan", "law", "government", "parliament", "executive", "judiciary", "rights", "bodies", "institutions", "citizenship"]):
        draw.line([(cx, cy - 35), (cx, cy + 35)], fill=border_color, width=4)
        draw.line([(cx - 20, cy + 35), (cx + 20, cy + 35)], fill=border_color, width=4)
        draw.line([(cx - 45, cy - 20), (cx + 45, cy - 20)], fill=border_color, width=4)
        
    elif any(k in title_lower for k in ["economy", "arthvyavastha", "economics", "budget", "finance", "fiscal", "money", "trade", "banking", "planning", "unemployment", "poverty"]):
        draw.line([(cx - 35, cy - 35), (cx - 35, cy + 35)], fill=border_color, width=3)
        draw.line([(cx - 35, cy + 35), (cx + 35, cy + 35)], fill=border_color, width=3)
        draw.line([(cx - 28, cy + 28), (cx - 12, cy + 12), (cx, cy + 20), (cx + 28, cy - 25)], fill=border_color, width=4)
        draw.polygon([(cx + 28, cy - 25), (cx + 14, cy - 25), (cx + 28, cy - 11)], fill=border_color)
        
    elif any(k in title_lower for k in ["science", "physics", "chemistry", "biology", "vigyan", "measurement", "motion", "force", "energy", "body", "health", "ecology", "environment", "plant", "animal", "genetics"]):
        draw.ellipse([(cx - 12, cy - 12), (cx + 12, cy + 12)], fill=border_color)
        draw.ellipse([(cx - 40, cy - 16), (cx + 40, cy + 16)], outline=border_color, width=2)
        draw.ellipse([(cx - 16, cy - 40), (cx + 16, cy + 40)], outline=border_color, width=2)
        
    elif any(k in title_lower for k in ["math", "mathematics", "ganit", "quantitative", "quant", "aptitude", "reasoning", "puzzle", "series", "matrix", "cube", "dice", "coding", "decoding", "relations"]):
        draw.polygon([(cx, cy - 35), (cx - 35, cy + 28), (cx + 35, cy + 28)], outline=border_color, width=2)
        
    else:
        draw.line([(cx, cy - 30), (cx, cy + 30)], fill=border_color, width=3)
        draw.polygon([(cx, cy - 30), (cx - 35, cy - 22), (cx - 35, cy + 22), (cx, cy + 30)], outline=border_color, width=2)
        draw.polygon([(cx, cy - 30), (cx + 35, cy - 22), (cx + 35, cy + 22), (cx, cy + 30)], outline=border_color, width=2)
        
    # Decorative diamond cluster
    for dx in [-30, 0, 30]:
        dcx = w // 2 + dx
        dcy = h // 2 - 120
        draw.polygon([(dcx, dcy - 7), (dcx + 7, dcy), (dcx, dcy + 7), (dcx - 7, dcy)], fill=border_color)
    
    # Text wrapping
    words = title.split()
    lines = []
    current_line = []
    for word in words:
        current_line.append(word)
        try:
            line_str = " ".join(current_line)
            bbox = draw.textbbox((0, 0), line_str, font=title_font)
            line_w = bbox[2] - bbox[0]
            if line_w > w - 300:
                current_line.pop()
                lines.append(" ".join(current_line))
                current_line = [word]
        except Exception:
            if len(current_line) > 4:
                current_line.pop()
                lines.append(" ".join(current_line))
                current_line = [word]
    if current_line:
        lines.append(" ".join(current_line))
        
    start_y = h // 2 - 15
    line_height = 65
    for i, line in enumerate(lines):
        y_pos = start_y + (i - len(lines)/2) * line_height
        draw.text((w//2 + 3, y_pos + 3), line, fill="#040811", font=title_font, anchor="mm")
        draw.text((w//2, y_pos), line, fill="#FFFFFF", font=title_font, anchor="mm")
        
    # Subtitle banner
    draw.text((w//2, h//2 + 140), subtitle.upper(), fill=border_color, font=sub_font, anchor="mm")
    
    filename = f"banner_{secrets.token_hex(8)}.jpg"
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads", "images")
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, filename)
    img.save(filepath, "JPEG", quality=92)
    
    # Upload to R2 if configured
    try:
        from app.services.r2_storage import upload_to_r2
        r2_key = f"classroom/banners/{filename}"
        r2_url = upload_to_r2(filepath, r2_key, "image/jpeg")
        if r2_url:
            return r2_url
    except Exception as r2_err:
        logger.error(f"Failed to upload premium banner to R2: {r2_err}")
        
    return f"/uploads/images/{filename}"



def generate_premium_image_locally(title: str, subtitle: str = "Subject") -> str:
    from PIL import Image, ImageDraw, ImageFont, ImageOps
    import os
    import secrets
    import httpx
    
    w, h = 512, 512
    img = None
    headers = {"User-Agent": "ClassroomImageGen/1.0 (admin@mr-ai.com)"}
    
    # 1. Search Wikimedia Commons for a relevant illustration matching the title
    wikimedia_url = fetch_wikimedia_image_url(title)
    if wikimedia_url:
        try:
            with httpx.Client(timeout=15.0) as client:
                resp = client.get(wikimedia_url, headers=headers)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    from io import BytesIO
                    raw_img = Image.open(BytesIO(resp.content))
                    # Crop and fit image centered
                    img = ImageOps.fit(raw_img, (w, h)).convert("RGB")
        except Exception as e:
            logger.warning(f"Failed to load Wikimedia image '{wikimedia_url}': {e}")
            
    # 2. Fallback to royal linear gradient background if download failed or no match found
    if img is None:
        img = Image.new("RGB", (w, h), "#0b192c")
        draw = ImageDraw.Draw(img)
        for y in range(h):
            r = int(11 + (30 - 11) * (y / h))
            g = int(25 + (62 - 25) * (y / h))
            b = int(44 + (98 - 44) * (y / h))
            draw.line([(0, y), (w, y)], fill=(r, g, b))
    else:
        # Create a beautiful dark blue transparent glassmorphic overlay for contrast
        overlay = Image.new("RGBA", (w, h), (11, 25, 44, 175)) # Navy overlay (approx 68% opacity)
        img.paste(overlay, (0, 0), overlay)
        
    draw = ImageDraw.Draw(img)
    border_color = "#D4AF37"  # Gold
    
    # Outer elegant frame
    draw.rectangle([(20, 20), (w - 20, h - 20)], outline=border_color, width=3)
    draw.rectangle([(26, 26), (w - 26, h - 26)], outline=border_color, width=1)
    
    # Corner ornaments
    draw.line([(12, 32), (42, 32)], fill=border_color, width=2)
    draw.line([(32, 12), (32, 42)], fill=border_color, width=2)
    
    draw.line([(w - 42, 32), (w - 12, 32)], fill=border_color, width=2)
    draw.line([(w - 32, 12), (w - 32, 42)], fill=border_color, width=2)
    
    draw.line([(12, h - 32), (42, h - 32)], fill=border_color, width=2)
    draw.line([(32, h - 42), (32, h - 12)], fill=border_color, width=2)
    
    draw.line([(w - 42, h - 32), (w - 12, h - 32)], fill=border_color, width=2)
    draw.line([(w - 32, h - 42), (w - 32, h - 12)], fill=border_color, width=2)
    
    # Font path
    font_path = "C:\\Windows\\Fonts\\georgiab.ttf"
    if not os.path.exists(font_path):
        font_path = "C:\\Windows\\Fonts\\timesbd.ttf"
    if not os.path.exists(font_path):
        font_path = "C:\\Windows\\Fonts\\arial.ttf"
        
    try:
        title_font = ImageFont.truetype(font_path, 34)
        sub_font = ImageFont.truetype(font_path, 18)
    except Exception:
        title_font = ImageFont.load_default()
        sub_font = ImageFont.load_default()
        
    # Central geometric medallion outline
    draw.ellipse([(w//2 - 90, h//2 - 120), (w//2 + 90, h//2 + 60)], outline=border_color, width=2)
    
    # Center coordinates of the medallion
    cx, cy = w // 2, h // 2 - 30
    title_lower = title.lower()
    
    # Draw keyword-specific gold vector icons
    if any(k in title_lower for k in ["history", "itihas", "ancient", "medieval", "modern", "war", "struggle", "movement", "period", "empire"]):
        draw.line([(cx, cy - 25), (cx, cy + 25)], fill=border_color, width=2) # Spine
        draw.polygon([(cx, cy - 25), (cx - 35, cy - 20), (cx - 35, cy + 20), (cx, cy + 25)], outline=border_color, width=2) # Left
        draw.polygon([(cx, cy - 25), (cx + 35, cy - 20), (cx + 35, cy + 20), (cx, cy + 25)], outline=border_color, width=2) # Right
        draw.line([(cx - 28, cy - 10), (cx - 8, cy - 7)], fill=border_color, width=1)
        draw.line([(cx - 28, cy), (cx - 8, cy + 3)], fill=border_color, width=1)
        draw.line([(cx - 28, cy + 10), (cx - 8, cy + 13)], fill=border_color, width=1)
        draw.line([(cx + 8, cy - 7), (cx + 28, cy - 10)], fill=border_color, width=1)
        draw.line([(cx + 8, cy + 3), (cx + 28, cy)], fill=border_color, width=1)
        draw.line([(cx + 8, cy + 13), (cx + 28, cy + 10)], fill=border_color, width=1)
        
    elif any(k in title_lower for k in ["geography", "bhugol", "climate", "resource", "earth", "settlement", "map", "space", "astronomy", "physical", "atmosphere"]):
        r = 30
        draw.ellipse([(cx - r, cy - r), (cx + r, cy + r)], outline=border_color, width=2)
        draw.line([(cx - r, cy), (cx + r, cy)], fill=border_color, width=1)
        draw.line([(cx, cy - r), (cx, cy + r)], fill=border_color, width=1)
        draw.ellipse([(cx - r, cy - r//2), (cx + r, cy + r//2)], outline=border_color, width=1)
        draw.ellipse([(cx - r//2, cy - r), (cx + r//2, cy + r)], outline=border_color, width=1)
        
    elif any(k in title_lower for k in ["polity", "constitution", "samvidhan", "law", "government", "parliament", "executive", "judiciary", "rights", "bodies", "institutions", "citizenship"]):
        draw.line([(cx, cy - 25), (cx, cy + 25)], fill=border_color, width=3)
        draw.line([(cx - 15, cy + 25), (cx + 15, cy + 25)], fill=border_color, width=3)
        draw.line([(cx - 35, cy - 15), (cx + 35, cy - 15)], fill=border_color, width=3)
        draw.line([(cx - 30, cy - 15), (cx - 40, cy + 5)], fill=border_color, width=1)
        draw.line([(cx - 30, cy - 15), (cx - 20, cy + 5)], fill=border_color, width=1)
        draw.arc([(cx - 40, cy + 3), (cx - 20, cy + 23)], start=0, end=180, fill=border_color, width=2)
        draw.line([(cx - 40, cy + 13), (cx - 20, cy + 13)], fill=border_color, width=2)
        draw.line([(cx + 30, cy - 15), (cx + 20, cy + 5)], fill=border_color, width=1)
        draw.line([(cx + 30, cy - 15), (cx + 40, cy + 5)], fill=border_color, width=1)
        draw.arc([(cx + 20, cy + 3), (cx + 40, cy + 23)], start=0, end=180, fill=border_color, width=2)
        draw.line([(cx + 20, cy + 13), (cx + 40, cy + 13)], fill=border_color, width=2)
        
    elif any(k in title_lower for k in ["economy", "arthvyavastha", "economics", "budget", "finance", "fiscal", "money", "trade", "banking", "planning", "unemployment", "poverty"]):
        draw.line([(cx - 25, cy - 25), (cx - 25, cy + 25)], fill=border_color, width=2)
        draw.line([(cx - 25, cy + 25), (cx + 25, cy + 25)], fill=border_color, width=2)
        draw.line([(cx - 20, cy + 20), (cx - 8, cy + 8), (cx, cy + 14), (cx + 20, cy - 18)], fill=border_color, width=3)
        draw.polygon([(cx + 20, cy - 18), (cx + 10, cy - 18), (cx + 20, cy - 8)], fill=border_color)
        
    elif any(k in title_lower for k in ["science", "physics", "chemistry", "biology", "vigyan", "measurement", "motion", "force", "energy", "body", "health", "ecology", "environment", "plant", "animal", "genetics"]):
        draw.ellipse([(cx - 8, cy - 8), (cx + 8, cy + 8)], fill=border_color)
        draw.ellipse([(cx - 30, cy - 12), (cx + 30, cy + 12)], outline=border_color, width=2)
        draw.ellipse([(cx - 12, cy - 30), (cx + 12, cy + 30)], outline=border_color, width=2)
        
    elif any(k in title_lower for k in ["math", "mathematics", "ganit", "quantitative", "quant", "aptitude", "reasoning", "puzzle", "series", "matrix", "cube", "dice", "coding", "decoding", "relations"]):
        draw.polygon([(cx, cy - 25), (cx - 25, cy + 20), (cx + 25, cy + 20)], outline=border_color, width=2)
        draw.ellipse([(cx - 12, cy - 4), (cx + 12, cy + 18)], outline=border_color, width=2)
        
    else:
        draw.line([(cx, cy - 20), (cx, cy + 20)], fill=border_color, width=2)
        draw.polygon([(cx, cy - 20), (cx - 25, cy - 15), (cx - 25, cy + 15), (cx, cy + 20)], outline=border_color, width=2)
        draw.polygon([(cx, cy - 20), (cx + 25, cy - 15), (cx + 25, cy + 15), (cx, cy + 20)], outline=border_color, width=2)
        
    # Decorative diamond cluster
    for dx in [-24, 0, 24]:
        dcx = w // 2 + dx
        dcy = h // 2 - 90
        draw.polygon([(dcx, dcy - 5), (dcx + 5, dcy), (dcx, dcy + 5), (dcx - 5, dcy)], fill=border_color)
    
    # Text wrapping
    words = title.split()
    lines = []
    current_line = []
    for word in words:
        current_line.append(word)
        try:
            line_str = " ".join(current_line)
            bbox = draw.textbbox((0, 0), line_str, font=title_font)
            line_w = bbox[2] - bbox[0]
            if line_w > w - 120:
                current_line.pop()
                lines.append(" ".join(current_line))
                current_line = [word]
        except Exception:
            if len(current_line) > 3:
                current_line.pop()
                lines.append(" ".join(current_line))
                current_line = [word]
    if current_line:
        lines.append(" ".join(current_line))
        
    start_y = h // 2 - 10
    line_height = 42
    for i, line in enumerate(lines):
        y_pos = start_y + (i - len(lines)/2) * line_height
        # Draw shadow
        draw.text((w//2 + 2, y_pos + 2), line, fill="#040811", font=title_font, anchor="mm")
        # Draw text
        draw.text((w//2, y_pos), line, fill="#FFFFFF", font=title_font, anchor="mm")
        
    # Subtitle banner
    draw.text((w//2, h//2 + 110), subtitle.upper(), fill=border_color, font=sub_font, anchor="mm")
    
    filename = f"classroom_{secrets.token_hex(8)}.jpg"
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads", "images")
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, filename)
    img.save(filepath, "JPEG", quality=92)
    
    # Upload to R2 if configured
    try:
        from app.services.r2_storage import upload_to_r2
        r2_key = f"classroom/images/{filename}"
        r2_url = upload_to_r2(filepath, r2_key, "image/jpeg")
        if r2_url:
            return r2_url
    except Exception as r2_err:
        logger.error(f"Failed to upload premium image to R2: {r2_err}")
        
    return f"/uploads/images/{filename}"



def download_and_save_image(image_url: str, fallback_keyword: str = None, subtitle: str = "Subject") -> str:
    import httpx
    try:
        filename = f"classroom_{secrets.token_hex(8)}.jpg"
        uploads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads", "images")
        os.makedirs(uploads_dir, exist_ok=True)
        filepath = os.path.join(uploads_dir, filename)
        
        with httpx.Client(timeout=10.0) as client:
            resp = client.get(image_url)
            if resp.status_code == 200 and len(resp.content) > 1000:  # Avoid error HTML/JSON responses
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                
                # Upload to R2 if configured
                try:
                    from app.services.r2_storage import upload_to_r2
                    r2_key = f"classroom/images/{filename}"
                    r2_url = upload_to_r2(filepath, r2_key, "image/jpeg")
                    if r2_url:
                        return r2_url
                except Exception as r2_err:
                    logger.error(f"Failed to upload downloaded image to R2: {r2_err}")
                    
                return f"/uploads/images/{filename}"
    except Exception as e:
        logger.warning(f"Error downloading image locally: {e}")
        
    # Fallback to local premium generation if Pollinations fails
    if fallback_keyword:
        try:
            return generate_premium_image_locally(fallback_keyword, subtitle)
        except Exception as ex:
            logger.error(f"Local premium image generation failed: {ex}")
            
    return image_url



class CreateTopicReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    video_length: Optional[int] = None
    script: Optional[str] = None
    image_url: Optional[str] = None

class CreateSubtopicReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = ""
    script: Optional[str] = None
    image_url: Optional[str] = None
    banner_url: Optional[str] = None


# ── Exam Endpoints ────────────────────────────────────────────────────────────

@router.get("/classroom/exams", tags=["Classroom"])
async def list_exams(client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exams = db.query(Exam).filter(Exam.client_id == client["client_id"]).order_by(Exam.created_at.desc()).all()
    return {"success": True, "exams": [e.to_dict() for e in exams]}


@router.post("/classroom/exams", tags=["Classroom"])
async def create_exam(req: CreateExamReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam_id = "exam-" + secrets.token_hex(8)
    exam = Exam(
        exam_id=exam_id,
        client_id=client["client_id"],
        name=req.name,
        category=req.category,
        image_url=req.image_url,
        description=req.description,
        created_at=datetime.utcnow()
    )
    db.add(exam)
    db.commit()
    db.refresh(exam)
    return {"success": True, "exam": exam.to_dict()}


@router.put("/classroom/exams/{exam_id}", tags=["Classroom"])
async def update_exam(exam_id: str, req: CreateExamReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found")
    exam.name = req.name
    exam.category = req.category
    exam.image_url = req.image_url
    exam.description = req.description
    db.commit()
    db.refresh(exam)
    return {"success": True, "exam": exam.to_dict()}


@router.delete("/classroom/exams/{exam_id}", tags=["Classroom"])
async def delete_exam(exam_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found")
    db.delete(exam)
    db.commit()
    return {"success": True, "message": "Exam deleted"}


@router.get("/classroom/exams/{exam_id}", tags=["Classroom"])
async def get_exam_details(exam_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).options(
        selectinload(Exam.papers)
        .selectinload(PaperClassroom.subjects)
        .selectinload(Subject.chapters)
        .selectinload(ChapterClassroom.topics)
        .selectinload(TopicClassroom.subtopics)
    ).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found")
    
    # Build hierarchical tree: Papers -> Subjects -> Chapters -> Topics -> Subtopics
    papers_data = []
    sorted_papers = sorted(exam.papers, key=lambda p: p.created_at or datetime.min, reverse=True)
    for paper in sorted_papers:
        p_dict = paper.to_dict()
        subjects_data = []
        for subject in paper.subjects:
            sub_dict = subject.to_dict()
            chapters_data = []
            for chapter in subject.chapters:
                ch_dict = chapter.to_dict()
                topics_data = []
                for topic in chapter.topics:
                    tp_dict = topic.to_dict()
                    subtopics_data = [st.to_dict() for st in topic.subtopics]
                    tp_dict["subtopics"] = subtopics_data
                    topics_data.append(tp_dict)
                ch_dict["topics"] = topics_data
                chapters_data.append(ch_dict)
            sub_dict["chapters"] = chapters_data
            subjects_data.append(sub_dict)
        p_dict["subjects"] = subjects_data
        papers_data.append(p_dict)
    
    return {
        "success": True,
        "exam": exam.to_dict(),
        "tree": papers_data
    }


# ── Paper Endpoints ───────────────────────────────────────────────────────────

@router.post("/classroom/exams/{exam_id}/papers", tags=["Classroom"])
async def create_paper(exam_id: str, req: CreatePaperReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found")
    paper_id = "paper-" + secrets.token_hex(8)
    paper = PaperClassroom(
        paper_id=paper_id,
        exam_id=exam_id,
        name=req.name,
        created_at=datetime.utcnow()
    )
    db.add(paper)
    db.commit()
    db.refresh(paper)
    return {"success": True, "paper": paper.to_dict()}


@router.put("/classroom/papers/{paper_id}", tags=["Classroom"])
async def update_paper(paper_id: str, req: CreatePaperReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    paper = db.query(PaperClassroom).join(Exam).filter(PaperClassroom.paper_id == paper_id, Exam.client_id == client["client_id"]).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")
    paper.name = req.name
    db.commit()
    db.refresh(paper)
    return {"success": True, "paper": paper.to_dict()}


@router.delete("/classroom/papers/{paper_id}", tags=["Classroom"])
async def delete_paper(paper_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    paper = db.query(PaperClassroom).join(Exam).filter(PaperClassroom.paper_id == paper_id, Exam.client_id == client["client_id"]).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")
    db.delete(paper)
    db.commit()
    return {"success": True, "message": "Paper deleted"}


# ── Subject Endpoints ─────────────────────────────────────────────────────────

@router.post("/classroom/papers/{paper_id}/subjects", tags=["Classroom"])
async def create_subject(paper_id: str, req: CreateSubjectReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    paper = db.query(PaperClassroom).join(Exam).filter(PaperClassroom.paper_id == paper_id, Exam.client_id == client["client_id"]).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")
    
    image_url = req.image_url
    if not image_url:
        exam_name = ""
        if paper.exam_id:
            exam = db.query(Exam).filter(Exam.exam_id == paper.exam_id).first()
            if exam:
                exam_name = exam.name
        context_str = f"{exam_name} {paper.name}".strip()
        image_url = await generate_ai_image_and_upload(req.name, subtitle="Subject", context=context_str)
    elif image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    subject_id = "subject-" + secrets.token_hex(8)
    subject = Subject(
        subject_id=subject_id,
        paper_id=paper_id,
        exam_id=paper.exam_id,
        name=req.name,
        color=req.color,
        image_url=image_url,
        created_at=datetime.utcnow()
    )
    db.add(subject)
    db.commit()
    db.refresh(subject)
    return {"success": True, "subject": subject.to_dict()}


@router.put("/classroom/subjects/{subject_id}", tags=["Classroom"])
async def update_subject(subject_id: str, req: CreateSubjectReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(Subject.subject_id == subject_id, Exam.client_id == client["client_id"]).first()
    if not subject:
        raise HTTPException(404, "Subject not found or access denied")
    
    image_url = req.image_url
    if image_url and image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    subject.name = req.name
    subject.color = req.color
    subject.image_url = image_url
    db.commit()
    db.refresh(subject)
    return {"success": True, "subject": subject.to_dict()}


@router.delete("/classroom/subjects/{subject_id}", tags=["Classroom"])
async def delete_subject(subject_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(Subject.subject_id == subject_id, Exam.client_id == client["client_id"]).first()
    if not subject:
        raise HTTPException(404, "Subject not found or access denied")
    db.delete(subject)
    db.commit()
    return {"success": True, "message": "Subject deleted"}


@router.post("/classroom/generate-image", tags=["Classroom"])
async def generate_classroom_image(req: GenerateImageReq, client: dict = Depends(_require_client)):
    if req.type == "subject":
        subtitle = "Subject"
        local_url = await generate_ai_image_and_upload(req.name, subtitle=subtitle, context=req.context)
    elif req.type == "current_affair":
        subtitle = "Current Affair"
        local_url = await generate_ai_image_and_upload(req.name, subtitle=subtitle, context=req.context)
    elif req.type == "topic":
        subtitle = "Topic"
        local_url = await generate_ai_image_and_upload(req.name, subtitle=subtitle, context=req.context)
    elif req.type == "subtopic":
        subtitle = "Subtopic"
        local_url = await generate_ai_image_and_upload(req.name, subtitle=subtitle, context=req.context)
    elif req.type == "subtopic_banner":
        local_url = await generate_banner_image_and_upload(req.name, context=req.context)
    else:
        subtitle = "Chapter"
        local_url = await generate_ai_image_and_upload(req.name, subtitle=subtitle, context=req.context)
    return {"success": True, "image_url": local_url}


@router.post("/classroom/generate-educational-prompt", tags=["Classroom"])
async def get_classroom_educational_prompt(req: GeneratePromptReq, client: dict = Depends(_require_client)):
    subtitle = req.type.capitalize()
    if req.type == "subtopic_banner":
        subtitle = "Subtopic Banner"
    elif req.type == "topic_banner":
        subtitle = "Topic Banner"    # treated as banner → 16:9 wide cinematic prompt
    elif req.type == "current_affair":
        subtitle = "Current Affair"
        
    prompt = await generate_educational_image_prompt(req.name, subtitle=subtitle, context=req.context)
    return {"success": True, "prompt": prompt}


def _crop_image_to_banner(img_bytes: bytes, out_path: str) -> bool:
    """Crop/resize a square (or any) image to 16:9 (1280x720) banner. Returns True on success."""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        w, h = img.size
        # Target 16:9
        target_w, target_h = 1280, 720
        target_ratio = target_w / target_h
        src_ratio = w / h
        if src_ratio > target_ratio:
            # Image is wider — crop sides
            new_w = int(h * target_ratio)
            x0 = (w - new_w) // 2
            img = img.crop((x0, 0, x0 + new_w, h))
        else:
            # Image is taller — crop top/bottom
            new_h = int(w / target_ratio)
            y0 = (h - new_h) // 2
            img = img.crop((0, y0, w, y0 + new_h))
        img = img.resize((target_w, target_h), Image.LANCZOS)
        img.save(out_path, "JPEG", quality=92)
        return True
    except Exception as e:
        logger.error(f"Banner crop failed: {e}")
        return False


async def _make_banner_from_url(image_url: str, entity_name: str, out_dir: str) -> str | None:
    """Download image from URL (local or remote), crop to 16:9, upload to R2. Returns final URL or None."""
    import httpx, secrets, os, shutil
    
    img_bytes = None
    
    # Try fetching image — handle both local paths and http URLs
    if image_url.startswith("http"):
        try:
            async with httpx.AsyncClient(timeout=20.0) as c:
                r = await c.get(image_url)
                if r.status_code == 200 and len(r.content) > 500:
                    img_bytes = r.content
        except Exception as e:
            logger.warning(f"Failed to fetch image from {image_url}: {e}")
    else:
        # Local file path like /uploads/images/abc.jpg
        local_path = os.path.join(os.getcwd(), image_url.lstrip("/").replace("/", os.sep))
        if os.path.exists(local_path):
            with open(local_path, "rb") as f:
                img_bytes = f.read()
    
    if not img_bytes:
        return None
    
    os.makedirs(out_dir, exist_ok=True)
    out_filename = f"banner_{secrets.token_hex(6)}.jpg"
    out_path = os.path.join(out_dir, out_filename)
    
    ok = _crop_image_to_banner(img_bytes, out_path)
    if not ok:
        return None
    
    # Upload to R2
    try:
        from app.services.r2_storage import upload_to_r2
        r2_key = f"classroom/banners/{out_filename}"
        r2_url = upload_to_r2(out_path, r2_key, "image/jpeg")
        if r2_url:
            return r2_url
    except Exception as e:
        logger.warning(f"R2 upload failed for banner: {e}")
    
    # Fallback: local URL
    return f"/uploads/images/{out_filename}"


@router.post("/classroom/exams/{exam_id}/generate-banners-from-images", tags=["Classroom"])
async def generate_banners_from_existing_images(
    exam_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    """
    Bulk-generate 16:9 banners for ALL subjects, chapters, topics, and subtopics
    under an exam that have an image_url but no banner_url.
    Crops the existing 1:1 image to 16:9 ratio — no new AI generation needed.
    """
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found or access denied")
    
    out_dir = os.path.join(os.getcwd(), "uploads", "images")
    
    stats = {"subjects": 0, "chapters": 0, "topics": 0, "subtopics": 0, "skipped": 0, "failed": 0}
    
    papers = db.query(PaperClassroom).join(Exam).filter(PaperClassroom.exam_id == exam_id).all()
    
    for paper in papers:
        subjects = db.query(Subject).filter(Subject.paper_id == paper.paper_id).all()
        for subject in subjects:
            # Subject banner
            if subject.image_url and not subject.banner_url:
                banner_url = await _make_banner_from_url(subject.image_url, subject.name, out_dir)
                if banner_url:
                    subject.banner_url = banner_url
                    stats["subjects"] += 1
                else:
                    stats["failed"] += 1
            elif subject.banner_url:
                stats["skipped"] += 1
            
            chapters = db.query(ChapterClassroom).filter(ChapterClassroom.subject_id == subject.subject_id).all()
            for chapter in chapters:
                # Chapter banner
                if chapter.image_url and not chapter.banner_url:
                    banner_url = await _make_banner_from_url(chapter.image_url, chapter.name, out_dir)
                    if banner_url:
                        chapter.banner_url = banner_url
                        stats["chapters"] += 1
                    else:
                        stats["failed"] += 1
                elif chapter.banner_url:
                    stats["skipped"] += 1
                
                topics = db.query(TopicClassroom).filter(TopicClassroom.chapter_id == chapter.chapter_id).all()
                for topic in topics:
                    # Topic banner
                    if topic.image_url and not topic.banner_url:
                        banner_url = await _make_banner_from_url(topic.image_url, topic.name, out_dir)
                        if banner_url:
                            topic.banner_url = banner_url
                            stats["topics"] += 1
                        else:
                            stats["failed"] += 1
                    elif topic.banner_url:
                        stats["skipped"] += 1
                    
                    subtopics = db.query(SubtopicClassroom).filter(SubtopicClassroom.topic_id == topic.topic_id).all()
                    for subtopic in subtopics:
                        # Subtopic banner
                        if subtopic.image_url and not subtopic.banner_url:
                            banner_url = await _make_banner_from_url(subtopic.image_url, subtopic.name, out_dir)
                            if banner_url:
                                subtopic.banner_url = banner_url
                                stats["subtopics"] += 1
                            else:
                                stats["failed"] += 1
                        elif subtopic.banner_url:
                            stats["skipped"] += 1
    
    db.commit()
    logger.info(f"Bulk banner generation complete for exam {exam_id}: {stats}")
    
    return {
        "success": True,
        "exam_id": exam_id,
        "generated": stats,
        "message": f"Banners generated: {stats['subjects']} subjects, {stats['chapters']} chapters, {stats['topics']} topics, {stats['subtopics']} subtopics. Skipped (already had banner): {stats['skipped']}. Failed: {stats['failed']}."
    }


@router.post("/classroom/regenerate-image", tags=["Classroom"])
async def regenerate_classroom_image(req: RegenerateImageReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    entity_id = req.entity_id
    entity_type = req.type.lower()
    
    if entity_type == "subject":
        subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(Subject.subject_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not subject:
            raise HTTPException(404, "Subject not found or access denied")
        exam_name = ""
        if subject.exam_id:
            exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
            if exam:
                exam_name = exam.name
        paper = db.query(PaperClassroom).filter(PaperClassroom.paper_id == subject.paper_id).first()
        paper_name = paper.name if paper else ""
        context_str = f"{exam_name} {paper_name}".strip()
        
        new_url = await generate_ai_image_and_upload(subject.name, subtitle="Subject", context=context_str)
        subject.image_url = new_url
        db.commit()
        db.refresh(subject)
        return {"success": True, "image_url": new_url, "entity": subject.to_dict()}
        
    elif entity_type == "chapter":
        chapter = db.query(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(ChapterClassroom.chapter_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not chapter:
            raise HTTPException(404, "Chapter not found or access denied")
        subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
        subject_name = subject.name if subject else ""
        exam_name = ""
        if subject and subject.exam_id:
            exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
            if exam:
                exam_name = exam.name
        context_str = f"{exam_name} {subject_name}".strip()
        
        new_url = await generate_ai_image_and_upload(chapter.name, subtitle="Chapter", context=context_str)
        chapter.image_url = new_url
        db.commit()
        db.refresh(chapter)
        return {"success": True, "image_url": new_url, "entity": chapter.to_dict()}
        
    elif entity_type == "topic":
        topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not topic:
            raise HTTPException(404, "Topic not found or access denied")
        chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
        chapter_name = chapter.name if chapter else ""
        subject_name = ""
        exam_name = ""
        if chapter:
            subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
            if subject:
                subject_name = subject.name
                if subject.exam_id:
                    exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                    if exam:
                        exam_name = exam.name
        context_str = f"{exam_name} {subject_name} {chapter_name}".strip()
        
        new_url = await generate_ai_image_and_upload(topic.name, subtitle="Topic", context=context_str)
        topic.image_url = new_url
        db.commit()
        db.refresh(topic)
        return {"success": True, "image_url": new_url, "entity": topic.to_dict()}
        
    elif entity_type == "subtopic":
        subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(SubtopicClassroom.subtopic_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not subtopic:
            raise HTTPException(404, "Subtopic not found or access denied")
        topic = db.query(TopicClassroom).filter(TopicClassroom.topic_id == subtopic.topic_id).first()
        topic_name = topic.name if topic else ""
        chapter_name = ""
        subject_name = ""
        exam_name = ""
        if topic:
            chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
            if chapter:
                chapter_name = chapter.name
                subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
                if subject:
                    subject_name = subject.name
                    if subject.exam_id:
                        exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                        if exam:
                            exam_name = exam.name
        context_str = f"{exam_name} {subject_name} {chapter_name} {topic_name}".strip()
        
        new_url = await generate_ai_image_and_upload(subtopic.name, subtitle="Subtopic", context=context_str)
        subtopic.image_url = new_url
        db.commit()
        db.refresh(subtopic)
        return {"success": True, "image_url": new_url, "entity": subtopic.to_dict()}
        
    elif entity_type == "subtopic_banner":
        subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(SubtopicClassroom.subtopic_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not subtopic:
            raise HTTPException(404, "Subtopic not found or access denied")
        topic = db.query(TopicClassroom).filter(TopicClassroom.topic_id == subtopic.topic_id).first()
        topic_name = topic.name if topic else ""
        chapter_name = ""
        subject_name = ""
        exam_name = ""
        if topic:
            chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
            if chapter:
                chapter_name = chapter.name
                subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
                if subject:
                    subject_name = subject.name
                    if subject.exam_id:
                        exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                        if exam:
                            exam_name = exam.name
        context_str = f"{exam_name} {subject_name} {chapter_name} {topic_name}".strip()
        
        new_url = await generate_banner_image_and_upload(subtopic.name, subtitle="Subtopic Banner", context=context_str)
        subtopic.banner_url = new_url
        db.commit()
        db.refresh(subtopic)
        return {"success": True, "banner_url": new_url, "entity": subtopic.to_dict()}
        
    elif entity_type == "topic_banner":
        topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == entity_id, Exam.client_id == client["client_id"]).first()
        if not topic:
            raise HTTPException(404, "Topic not found or access denied")
        chapter_name = ""
        subject_name = ""
        exam_name = ""
        chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
        if chapter:
            chapter_name = chapter.name
            subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
            if subject:
                subject_name = subject.name
                if subject.exam_id:
                    exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                    if exam:
                        exam_name = exam.name
        context_str = f"{exam_name} {subject_name} {chapter_name}".strip()
        
        new_url = await generate_banner_image_and_upload(topic.name, subtitle="Topic Banner", context=context_str)
        topic.banner_url = new_url
        db.commit()
        db.refresh(topic)
        return {"success": True, "banner_url": new_url, "entity": topic.to_dict()}
        
    else:
        raise HTTPException(400, "Invalid entity type for image regeneration")


# ── Chapter Endpoints ─────────────────────────────────────────────────────────

@router.post("/classroom/subjects/{subject_id}/chapters", tags=["Classroom"])
async def create_chapter(subject_id: str, req: CreateChapterReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(Subject.subject_id == subject_id, Exam.client_id == client["client_id"]).first()
    if not subject:
        raise HTTPException(404, "Subject not found or access denied")
        
    image_url = req.image_url
    if not image_url:
        exam_name = ""
        if subject.exam_id:
            exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
            if exam:
                exam_name = exam.name
        context_str = f"{exam_name} {subject.name}".strip()
        image_url = await generate_ai_image_and_upload(req.name, subtitle="Chapter", context=context_str)
    elif image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    chapter_id = "chapter-" + secrets.token_hex(8)
    chapter = ChapterClassroom(
        chapter_id=chapter_id,
        subject_id=subject_id,
        name=req.name,
        image_url=image_url,
        created_at=datetime.utcnow()
    )
    db.add(chapter)
    db.commit()
    db.refresh(chapter)
    return {"success": True, "chapter": chapter.to_dict()}


@router.put("/classroom/chapters/{chapter_id}", tags=["Classroom"])
async def update_chapter(chapter_id: str, req: CreateChapterReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    chapter = db.query(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(ChapterClassroom.chapter_id == chapter_id, Exam.client_id == client["client_id"]).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found or access denied")
    
    image_url = req.image_url
    if image_url and image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    chapter.name = req.name
    chapter.image_url = image_url
    db.commit()
    db.refresh(chapter)
    return {"success": True, "chapter": chapter.to_dict()}


@router.delete("/classroom/chapters/{chapter_id}", tags=["Classroom"])
async def delete_chapter(chapter_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    chapter = db.query(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(ChapterClassroom.chapter_id == chapter_id, Exam.client_id == client["client_id"]).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found or access denied")
    db.delete(chapter)
    db.commit()
    return {"success": True, "message": "Chapter deleted"}


# ── Topic Endpoints ───────────────────────────────────────────────────────────

@router.post("/classroom/chapters/{chapter_id}/topics", tags=["Classroom"])
async def create_topic(chapter_id: str, req: CreateTopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    chapter = db.query(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(ChapterClassroom.chapter_id == chapter_id, Exam.client_id == client["client_id"]).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found or access denied")
    
    image_url = req.image_url
    if not image_url:
        subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
        subject_name = subject.name if subject else ""
        exam_name = ""
        if subject and subject.exam_id:
            exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
            if exam:
                exam_name = exam.name
        context_str = f"{exam_name} {subject_name} {chapter.name}".strip()
        image_url = await generate_ai_image_and_upload(req.name, subtitle="Topic", context=context_str)
    elif image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    topic_id = "topic-" + secrets.token_hex(8)
    topic = TopicClassroom(
        topic_id=topic_id,
        chapter_id=chapter_id,
        name=req.name,
        video_length=req.video_length,
        script=req.script,
        image_url=image_url,
        created_at=datetime.utcnow()
    )
    db.add(topic)
    db.commit()
    db.refresh(topic)
    return {"success": True, "topic": topic.to_dict()}


@router.put("/classroom/topics/{topic_id}", tags=["Classroom"])
async def update_topic(topic_id: str, req: CreateTopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == topic_id, Exam.client_id == client["client_id"]).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
    topic.name = req.name
    topic.video_length = req.video_length
    topic.script = req.script
    if req.image_url:
        if req.image_url.startswith("http") and not req.image_url.startswith("/uploads"):
            topic.image_url = download_and_save_image(req.image_url, fallback_keyword=req.name)
        else:
            topic.image_url = req.image_url
    db.commit()
    db.refresh(topic)
    return {"success": True, "topic": topic.to_dict()}


@router.delete("/classroom/topics/{topic_id}", tags=["Classroom"])
async def delete_topic(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == topic_id, Exam.client_id == client["client_id"]).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
    db.delete(topic)
    db.commit()
    return {"success": True, "message": "Topic deleted"}


# ── Subtopic Endpoints ────────────────────────────────────────────────────────

@router.post("/classroom/topics/{topic_id}/subtopics", tags=["Classroom"])
async def create_subtopic(topic_id: str, req: CreateSubtopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == topic_id, Exam.client_id == client["client_id"]).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
    
    # Build context
    chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
    chapter_name = chapter.name if chapter else ""
    subject_name = ""
    exam_name = ""
    if chapter:
        subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
        if subject:
            subject_name = subject.name
            if subject.exam_id:
                exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                if exam:
                    exam_name = exam.name
    context_str = f"{exam_name} {subject_name} {chapter_name} {topic.name}".strip()
    
    image_url = req.image_url
    if not image_url:
        image_url = await generate_ai_image_and_upload(req.name, subtitle="Subtopic", context=context_str)
    elif image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)
        
    banner_url = req.banner_url
    if not banner_url:
        banner_url = await generate_banner_image_and_upload(req.name, context=context_str)
    elif banner_url.startswith("http") and not banner_url.startswith("/uploads"):
        banner_url = download_and_save_image(banner_url, fallback_keyword=req.name)
        
    subtopic_id = "subtopic-" + secrets.token_hex(8)
    subtopic = SubtopicClassroom(
        subtopic_id=subtopic_id,
        topic_id=topic_id,
        name=req.name,
        description=req.description,
        script=req.script,
        image_url=image_url,
        banner_url=banner_url,
        created_at=datetime.utcnow()
    )
    db.add(subtopic)
    db.commit()
    db.refresh(subtopic)
    return {"success": True, "subtopic": subtopic.to_dict()}


@router.put("/classroom/subtopics/{subtopic_id}", tags=["Classroom"])
async def update_subtopic(subtopic_id: str, req: CreateSubtopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(SubtopicClassroom.subtopic_id == subtopic_id, Exam.client_id == client["client_id"]).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
    subtopic.name = req.name
    subtopic.description = req.description
    if req.script is not None:
        subtopic.script = req.script
    if req.image_url:
        if req.image_url.startswith("http") and not req.image_url.startswith("/uploads"):
            subtopic.image_url = download_and_save_image(req.image_url, fallback_keyword=req.name)
        else:
            subtopic.image_url = req.image_url
    if req.banner_url:
        if req.banner_url.startswith("http") and not req.banner_url.startswith("/uploads"):
            subtopic.banner_url = download_and_save_image(req.banner_url, fallback_keyword=req.name)
        else:
            subtopic.banner_url = req.banner_url
    db.commit()
    db.refresh(subtopic)
    return {"success": True, "subtopic": subtopic.to_dict()}


@router.delete("/classroom/subtopics/{subtopic_id}", tags=["Classroom"])
async def delete_subtopic(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(SubtopicClassroom.subtopic_id == subtopic_id, Exam.client_id == client["client_id"]).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
    db.delete(subtopic)
    db.commit()
    return {"success": True, "message": "Subtopic deleted"}


@router.post("/classroom/subtopics/{subtopic_id}/generate-description", tags=["Classroom"])
async def generate_subtopic_description(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id, 
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
        
    topic = subtopic.topic
    chapter = topic.chapter
    subject = chapter.subject
    
    prompt = f"""Subject: {subject.name}, Chapter: {chapter.name}, Topic: {topic.name}.
Generate a detailed, premium study description and conceptual explanation for the subtopic: "{subtopic.name}".
Write a comprehensive explanation formatted beautifully with Markdown:
- Use heading hierarchy (e.g. ### Key Concepts, ### Practical Examples, etc.)
- Explain terms clearly.
- Provide examples/illustrations.
Return ONLY the formatted markdown text. Do not wrap in a markdown block like ```markdown, do not add metadata, just return the text."""

    try:
        from app.services.llm import generate_simple_response
        desc_text = await generate_simple_response(prompt, system_prompt="You are an expert educator who writes highly informative, clean study materials.")
        subtopic.description = desc_text
        db.commit()
        db.refresh(subtopic)
        return {"success": True, "description": desc_text, "subtopic": subtopic.to_dict()}
    except Exception as e:
        logger.error(f"Error generating subtopic description: {e}")
        raise HTTPException(500, f"Failed to generate description: {str(e)}")


@router.post("/classroom/topics/{topic_id}/generate-description", tags=["Classroom"])
async def generate_topic_description(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id, 
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
        
    chapter = topic.chapter
    subject = chapter.subject if chapter else None
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    
    prompt = f"""Subject: {subject_name}, Chapter: {chapter_name}.
Generate a detailed, premium study description and conceptual explanation for the topic: "{topic.name}".
Write a comprehensive explanation formatted beautifully with Markdown:
- Use heading hierarchy (e.g. ### Key Concepts, ### Practical Examples, etc.)
- Explain terms clearly.
- Provide examples/illustrations.
Return ONLY the formatted markdown text. Do not wrap in a markdown block like ```markdown, do not add metadata, just return the text."""

    try:
        from app.services.llm import generate_simple_response
        desc_text = await generate_simple_response(prompt, system_prompt="You are an expert educator who writes highly informative, clean study materials.")
        topic.description = desc_text
        db.commit()
        db.refresh(topic)
        return {"success": True, "description": desc_text, "topic": topic.to_dict()}
    except Exception as e:
        logger.error(f"Error generating topic description: {e}")
        raise HTTPException(500, f"Failed to generate description: {str(e)}")


import urllib.parse
import re

def process_markdown_images(text: str) -> str:
    if not text:
        return text
    # Pattern to find [IMAGE: descriptive prompt]
    pattern = re.compile(r'\[IMAGE:\s*([^\]]+)\]')
    def replacer(match):
        prompt = match.group(1).strip()
        encoded = urllib.parse.quote(prompt)
        url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&model=flux"
        return f"![{prompt}]({url})"
    return pattern.sub(replacer, text)


@router.post("/classroom/subtopics/{subtopic_id}/generate-notes", tags=["Classroom"])
async def generate_subtopic_notes(subtopic_id: str, language: str = Form("English"), client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id, 
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
        
    topic = subtopic.topic
    chapter = topic.chapter
    subject = chapter.subject
    
    prompt = f"""Subject: {subject.name}, Chapter: {chapter.name}, Topic: {topic.name}.
Generate extremely detailed, comprehensive, and structured study notes for the subtopic: "{subtopic.name}".
Language: {language}. Write all explanations, headings, and content in {language}.
Provide exhaustive coverage of all concepts, core definitions, underlying principles, step-by-step explanations, math formulas, code snippets, key takeaways, and practical examples.
Do not summarize briefly; ensure that every important aspect is fully explained in depth so a student can learn the subtopic thoroughly from these notes.

Formatting requirements:
1. Use ONLY pure Markdown — no HTML tags, no <div>, no inline styles.
2. Use blockquotes for callouts:
   - Tips:       > 💡 **Tip:** ...
   - Warnings:   > ⚠️ **Warning:** ...
   - Key Points: > ✅ **Key Point:** ...
   - Examples:   > 📌 **Example:** ...
3. Include at least 2-3 high-quality educational diagrams using the tag `[IMAGE: descriptive prompt for educational diagram]`.
4. Structure sections:
   # 📖 Introduction & Concept Overview
   # 📌 Key Formulas, Rules & Syntax
   # 💡 In-Depth Analysis & Practical Examples
   # ⚡ Critical Takeaways & Common Pitfalls
   # 🔍 Quick Revision & Summary Points

Return ONLY the formatted markdown text. Do not wrap in ```markdown blocks."""

    try:
        from app.services.llm import generate_simple_response
        notes_text = await generate_simple_response(prompt, system_prompt="You are an expert academic tutor specializing in writing comprehensive and extremely high-quality, in-depth study notes.")
        processed_notes = process_markdown_images(notes_text)
        subtopic.notes = processed_notes
        db.commit()
        db.refresh(subtopic)
        return {"success": True, "notes": processed_notes, "markdown": processed_notes, "subtopic": subtopic.to_dict()}
    except Exception as e:
        logger.error(f"Error generating subtopic notes: {e}")
        raise HTTPException(500, f"Failed to generate notes: {str(e)}")


@router.post("/classroom/topics/{topic_id}/generate-notes", tags=["Classroom"])
async def generate_topic_notes(topic_id: str, language: str = Form("English"), client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id, 
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
        
    chapter = topic.chapter
    subject = chapter.subject if chapter else None
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    
    prompt = f"""Subject: {subject_name}, Chapter: {chapter_name}.
Generate extremely detailed, comprehensive, and structured study notes for the topic: "{topic.name}".
Language: {language}. Write all explanations, headings, and content in {language}.
Provide exhaustive coverage of all concepts, core definitions, underlying principles, step-by-step explanations, math formulas, code snippets, key takeaways, and practical examples.
Do not summarize briefly; ensure that every important aspect is fully explained in depth so a student can learn the topic thoroughly from these notes.

Formatting requirements:
1. Use ONLY pure Markdown — no HTML tags, no <div>, no inline styles.
2. Use blockquotes for callouts:
   - Tips:       > 💡 **Tip:** ...
   - Warnings:   > ⚠️ **Warning:** ...
   - Key Points: > ✅ **Key Point:** ...
   - Examples:   > 📌 **Example:** ...
3. Include at least 2-3 high-quality educational diagrams using the tag `[IMAGE: descriptive prompt for educational diagram]`.
4. Structure sections:
   # 📖 Introduction & Concept Overview
   # 📌 Key Formulas, Rules & Syntax
   # 💡 In-Depth Analysis & Practical Examples
   # ⚡ Critical Takeaways & Common Pitfalls
   # 🔍 Quick Revision & Summary Points

Return ONLY the formatted markdown text. Do not wrap in ```markdown blocks."""

    try:
        from app.services.llm import generate_simple_response
        notes_text = await generate_simple_response(prompt, system_prompt="You are an expert academic tutor specializing in writing comprehensive and extremely high-quality, in-depth study notes.")
        processed_notes = process_markdown_images(notes_text)
        topic.notes = processed_notes
        db.commit()
        db.refresh(topic)
        return {"success": True, "notes": processed_notes, "markdown": processed_notes, "topic": topic.to_dict()}
    except Exception as e:
        logger.error(f"Error generating topic notes: {e}")
        raise HTTPException(500, f"Failed to generate notes: {str(e)}")


@router.get("/classroom/subtopics/{subtopic_id}/download-notes-pdf", tags=["Classroom"])
async def download_subtopic_notes_pdf(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")

    if not subtopic.notes and not subtopic.description:
        raise HTTPException(400, "No notes or description have been generated for this subtopic yet.")

    topic = subtopic.topic
    chapter = topic.chapter if topic else None
    subject = chapter.subject if chapter else None
    paper = subject.paper if subject else None
    exam = paper.exam if paper else None

    exam_name = exam.name if exam else "General Exam"
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name if topic else "General"
    subtopic_name = subtopic.name

    try:
        from app.services.pdf_generator import generate_notes_pdf_bytes
        from fastapi.responses import Response

        pdf_bytes = generate_notes_pdf_bytes(
            subtopic_name=subtopic_name,
            topic_name=topic_name,
            chapter_name=chapter_name,
            subject_name=subject_name,
            exam_name=exam_name,
            description_text=subtopic.description or "",
            notes_text=subtopic.notes or "",
        )

        safe_filename = "".join(
            [c if c.isalnum() or c in ("_", ".") else "_" for c in f"{subtopic_name}_notes.pdf"]
        )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )
    except Exception as e:
        logger.error(f"Error generating notes PDF: {e}")
        raise HTTPException(500, f"Failed to generate PDF: {str(e)}")


@router.get("/classroom/topics/{topic_id}/download-notes-pdf", tags=["Classroom"])
async def download_topic_notes_pdf(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")

    if not topic.notes and not topic.description:
        raise HTTPException(400, "No notes or description have been generated for this topic yet.")

    chapter = topic.chapter
    subject = chapter.subject if chapter else None
    paper = subject.paper if subject else None
    exam = paper.exam if paper else None

    exam_name = exam.name if exam else "General Exam"
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name

    try:
        from app.services.pdf_generator import generate_notes_pdf_bytes
        from fastapi.responses import Response

        pdf_bytes = generate_notes_pdf_bytes(
            subtopic_name=topic_name,
            topic_name=topic_name,
            chapter_name=chapter_name,
            subject_name=subject_name,
            exam_name=exam_name,
            description_text=topic.description or "",
            notes_text=topic.notes or "",
        )

        safe_filename = "".join(
            [c if c.isalnum() or c in ("_", ".") else "_" for c in f"{topic_name}_notes.pdf"]
        )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )
    except Exception as e:
        logger.error(f"Error generating topic notes PDF: {e}")
        raise HTTPException(500, f"Failed to generate PDF: {str(e)}")


# ── Subtopic Reels History Endpoint ─────────────────────────────────────────────

@router.get("/classroom/subtopics/{subtopic_id}/reels", tags=["Classroom"])
async def get_subtopic_reels(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found")
    
    from app.core.models import SocialContent
    reels = db.query(SocialContent).filter(
        SocialContent.client_id == client["client_id"],
        SocialContent.content_type == "reel"
    ).order_by(SocialContent.created_at.desc()).all()
    
    subtopic_reels = []
    for r in reels:
        try:
            meta = r.metadata_info
            if meta and meta.get("subtopic_id") == subtopic_id:
                subtopic_reels.append(r.to_dict())
        except Exception as e:
            logger.warning(f"Error parsing metadata for social content {r.content_id}: {e}")
    
    return {"success": True, "reels": subtopic_reels}


@router.get("/classroom/topics/{topic_id}/reels", tags=["Classroom"])
async def get_topic_reels(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
        
    subtopic_ids = [st.subtopic_id for st in topic.subtopics]
    
    from app.core.models import SocialContent
    reels = db.query(SocialContent).filter(
        SocialContent.client_id == client["client_id"],
        SocialContent.content_type == "reel"
    ).order_by(SocialContent.created_at.desc()).all()
    
    topic_reels = []
    for r in reels:
        try:
            meta = r.metadata_info
            if meta:
                # Direct topic match OR matches one of the topic's subtopics
                r_subtopic_id = meta.get("subtopic_id")
                r_topic_id = meta.get("topic_id")
                if r_topic_id == topic_id or (r_subtopic_id and r_subtopic_id in subtopic_ids):
                    topic_reels.append(r.to_dict())
        except Exception as e:
            logger.warning(f"Error parsing metadata for social content {r.content_id}: {e}")
            
    return {"success": True, "reels": topic_reels}


# ── Public Subtopic Reels Endpoint (No Auth Needed for other projects) ──

@router.get("/classroom/public/subtopics/{subtopic_id}/reels", tags=["Classroom Public"])
async def get_public_subtopic_reels(subtopic_id: str, db: Session = Depends(get_db)):
    """
    Publicly fetch all generated reels for a specific subtopic.
    Requires no X-App-Token, making it extremely easy for other projects to consume.
    """
    subtopic = db.query(SubtopicClassroom).filter(SubtopicClassroom.subtopic_id == subtopic_id).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found")
        
    from app.core.models import SocialContent
    
    # Fetch all reels without filtering by client_id to allow simple cross-project integration
    reels = db.query(SocialContent).filter(
        SocialContent.content_type == "reel"
    ).order_by(SocialContent.created_at.desc()).all()
    
    subtopic_reels = []
    for r in reels:
        try:
            meta = r.metadata_info
            if meta and meta.get("subtopic_id") == subtopic_id:
                subtopic_reels.append(r.to_dict())
        except Exception as e:
            logger.warning(f"Error parsing metadata for social content {r.content_id}: {e}")
            
    return {"success": True, "reels": subtopic_reels}


# ── Exam History (Reels) Endpoint ─────────────────────────────────────────────

@router.get("/classroom/exams/{exam_id}/history", tags=["Classroom"])
async def get_exam_history(exam_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found")
        
    from app.core.models import SocialContent
    
    # Query client's reels
    reels = db.query(SocialContent).filter(
        SocialContent.client_id == client["client_id"],
        SocialContent.content_type == "reel"
    ).order_by(SocialContent.created_at.desc()).all()
    
    # Filter by exam_id inside metadata_json
    exam_reels = []
    for r in reels:
        try:
            meta = r.metadata_info
            if meta and meta.get("exam_id") == exam_id:
                exam_reels.append(r.to_dict())
        except Exception as e:
            logger.warning(f"Error parsing metadata for social content {r.content_id}: {e}")
            
    return {"success": True, "reels": exam_reels}

# ── Quiz Generation Endpoint ──────────────────────────────────────────────────

@router.post("/classroom/subtopics/{subtopic_id}/quiz/generate", tags=["Classroom"])
async def generate_subtopic_quiz(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
        
    topic = subtopic.topic
    chapter = topic.chapter if topic else None
    subject = chapter.subject if chapter else None
    paper = subject.paper if subject else None
    exam = paper.exam if paper else None
    
    exam_name = exam.name if exam else "General Exam"
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name if topic else "General"
    subtopic_name = subtopic.name
    
    study_material = subtopic.description or subtopic.notes or subtopic.name
    
    prompt = f"""You are an expert academic examiner creating exam questions for the exam: "{exam_name}".
Generate exactly 5 multiple-choice questions based on the provided study material.
The questions must be highly professional and tailored specifically to the standards and difficulty level of the "{exam_name}" exam.
Subject: {subject_name}
Chapter: {chapter_name}
Topic: {topic_name}
Subtopic: {subtopic_name}

Return the response as a strict JSON array.
Format example:
[
  {{
    "question": "What is the capital of France?",
    "options": ["Paris", "London", "Berlin", "Rome"],
    "answer": "Paris"
  }}
]
IMPORTANT: Return ONLY the JSON array, no markdown blocks, no extra text.
"""
    try:
        from app.services.llm import generate_answer
        import json
        import re
        
        raw_response = await generate_answer(prompt, study_material)
        
        # Clean markdown code blocks if any
        clean_json = re.sub(r'```json\s*', '', raw_response)
        clean_json = re.sub(r'\s*```', '', clean_json).strip()
        
        quiz_data = json.loads(clean_json)
        if not isinstance(quiz_data, list):
            raise ValueError("Response is not a JSON array")
            
        return {"success": True, "quiz": quiz_data}
    except Exception as e:
        logger.error(f"Error generating quiz: {e}")
        raise HTTPException(500, f"Failed to generate quiz: {str(e)}")


@router.post("/classroom/topics/{topic_id}/quiz/generate", tags=["Classroom"])
async def generate_topic_quiz(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
        
    chapter = topic.chapter if topic else None
    subject = chapter.subject if chapter else None
    paper = subject.paper if subject else None
    exam = paper.exam if paper else None
    
    exam_name = exam.name if exam else "General Exam"
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name
    
    study_material = topic.description or topic.notes or topic.name
    
    prompt = f"""You are an expert academic examiner creating exam questions for the exam: "{exam_name}".
Generate exactly 5 multiple-choice questions based on the provided study material.
The questions must be highly professional and tailored specifically to the standards and difficulty level of the "{exam_name}" exam.
Subject: {subject_name}
Chapter: {chapter_name}
Topic: {topic_name}

Return the response as a strict JSON array.
Format example:
[
  {{
    "question": "What is the capital of France?",
    "options": ["Paris", "London", "Berlin", "Rome"],
    "answer": "Paris"
  }}
]
IMPORTANT: Return ONLY the JSON array, no markdown blocks, no extra text.
"""
    try:
        from app.services.llm import generate_answer
        import json
        import re
        
        raw_response = await generate_answer(prompt, study_material)
        
        # Clean markdown code blocks if any
        clean_json = re.sub(r'```json\s*', '', raw_response)
        clean_json = re.sub(r'\s*```', '', clean_json).strip()
        
        quiz_data = json.loads(clean_json)
        if not isinstance(quiz_data, list):
            raise ValueError("Response is not a JSON array")
            
        return {"success": True, "quiz": quiz_data}
    except Exception as e:
        logger.error(f"Error generating topic quiz: {e}")
        raise HTTPException(500, f"Failed to generate quiz: {str(e)}")

# ── Bulk Auto-Generate Course Endpoint ────────────────────────────────────────

async def run_generation_pipeline(paper_id: str, exam_id: str, exam_category: str, exam_name: str, paper_name: str):
    from app.core.database import get_session_local
    from app.core.config import settings
    from openai import AsyncOpenAI
    import json
    import re
    import secrets
    from datetime import datetime
    import asyncio
    
    SessionLocal = get_session_local()
    db = SessionLocal()
    
    # ── Use Groq API (Free: 30 RPM, 1000 RPD, 100K TPD) ──
    groq_key = settings.GROQ_API_KEY
    groq_model = settings.GROQ_MODEL or "llama-3.3-70b-versatile"
    
    groq_client = AsyncOpenAI(
        api_key=groq_key,
        base_url="https://api.groq.com/openai/v1",
        timeout=120.0
    )
    
    call_count = 0  # Track total API calls
    
    async def ask_llm(prompt_text, max_retries=3):
        """Call Groq with retry on 429 rate-limit errors."""
        nonlocal call_count
        
        for attempt in range(max_retries):
            try:
                response = await groq_client.chat.completions.create(
                    model=groq_model,
                    messages=[
                        {"role": "system", "content": "You are a specialized curriculum generator. Return strictly valid JSON containing exactly what is requested, no markdown wrappers, no extra text."},
                        {"role": "user", "content": prompt_text}
                    ],
                    max_tokens=2000,
                    temperature=0.7
                )
                raw = response.choices[0].message.content
                clean = re.sub(r'```json\s*', '', raw, flags=re.IGNORECASE)
                clean = re.sub(r'\s*```', '', clean).strip()
                call_count += 1
                logger.info(f"📡 Groq call #{call_count} success (attempt {attempt+1})")
                return json.loads(clean)
            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "rate_limit" in err_str.lower():
                    # Default fallback wait time
                    wait_time = (attempt + 1) * 20
                    
                    # Try to extract the exact wait time from response headers if available
                    try:
                        if hasattr(e, "response") and e.response is not None:
                            retry_after = e.response.headers.get("retry-after") or e.response.headers.get("x-ratelimit-reset")
                            if retry_after:
                                if retry_after.isdigit():
                                    wait_time = int(retry_after) + 2
                                else:
                                    # Parse duration like "1m55s" or similar
                                    m_match = re.search(r"(\d+)m", retry_after)
                                    s_match = re.search(r"(\d+)s", retry_after)
                                    parsed_time = 0
                                    if m_match:
                                        parsed_time += int(m_match.group(1)) * 60
                                    if s_match:
                                        parsed_time += int(s_match.group(1))
                                    if parsed_time > 0:
                                        wait_time = parsed_time + 2
                    except Exception as parse_err:
                        logger.debug(f"Failed to parse retry-after header: {parse_err}")
                    
                    # If headers didn't work, try to extract wait time from the exception message using regex
                    # e.g., "Please try again in 155.084s." or "try again in 155 seconds"
                    try:
                        match = re.search(r"try again in (\d+(?:\.\d+)?)\s*s", err_str)
                        if not match:
                            match = re.search(r"try again in (\d+(?:\.\d+)?)\s*second", err_str)
                        if match:
                            wait_time = int(float(match.group(1))) + 2
                    except Exception as regex_err:
                        logger.debug(f"Failed to parse wait time via regex: {regex_err}")
                        
                    logger.warning(f"⚠️ Groq 429 rate limit hit. Waiting {wait_time}s... (attempt {attempt+1}/{max_retries})")
                    await asyncio.sleep(wait_time)
                else:
                    logger.error(f"❌ Groq/JSON Error: {e}")
                    return None
        
        logger.error(f"❌ Max retries exhausted for prompt")
        return None

    try:
        logger.info(f"🚀 Started Groq Syllabus Pipeline for '{paper_name}' (model: {groq_model})")
        
        # ── STEP 1: Generate Subjects ──
        prompt_subjs = f"""Exam: {exam_name} ({exam_category}), Paper: {paper_name}. 
Generate the standard subjects (5-10 max) for this paper. 
Return JSON: {{"subjects": [{{"name": "Subject Name", "color": "#HEX"}}]}}"""
        
        res_subjs = await ask_llm(prompt_subjs)
        if not res_subjs or 'subjects' not in res_subjs:
            logger.error("Failed to generate subjects. Aborting.")
            return
            
        subject_records = []
        for s in res_subjs['subjects']:
            subj_id = "subject-" + secrets.token_hex(8)
            new_subject = Subject(subject_id=subj_id, exam_id=exam_id, paper_id=paper_id, name=s['name'], color=s.get('color', '#4f46e5'))
            db.add(new_subject)
            subject_records.append((subj_id, s['name']))
        db.commit()
        logger.info(f"✅ Step 1 done: {len(subject_records)} subjects created")
        await asyncio.sleep(3)
        
        # ── STEP 2: Generate Chapters (per subject) ──
        all_chapters = []
        for subj_id, subj_name in subject_records:
            prompt = f"""Exam: {exam_name}, Subject: {subj_name}. 
Generate comprehensive chapters (5-15 max). 
Return JSON: {{"chapters": [{{"name": "Chapter Name"}}]}}"""
            
            res = await ask_llm(prompt)
            if res and 'chapters' in res:
                for c in res['chapters']:
                    chap_id = "chapter-" + secrets.token_hex(8)
                    db.add(ChapterClassroom(chapter_id=chap_id, subject_id=subj_id, name=c['name']))
                    all_chapters.append((chap_id, c['name'], subj_name))
                db.commit()
            await asyncio.sleep(3)  # Respect 30 RPM
        
        logger.info(f"✅ Step 2 done: {len(all_chapters)} chapters created")
        
        # ── STEP 3: Generate Topics + Subtopics together (saves API calls) ──
        all_topics = []
        all_subtopics = []
        for chap_id, cname, sname in all_chapters:
            prompt = f"""Subject: {sname}, Chapter: {cname}. 
Generate all topics (3-8) with their subtopics (2-4 each).
Return JSON: {{"topics": [{{"name": "Topic Name", "subtopics": [{{"name": "Subtopic Name"}}]}}]}}"""
            
            res = await ask_llm(prompt)
            if res and 'topics' in res:
                for t in res['topics']:
                    top_id = "topic-" + secrets.token_hex(8)
                    db.add(TopicClassroom(topic_id=top_id, chapter_id=chap_id, name=t['name']))
                    all_topics.append((top_id, t['name'], cname, sname))
                    
                    # Also create subtopics from the same response
                    for st in t.get('subtopics', []):
                        sub_id = "subtopic-" + secrets.token_hex(8)
                        db.add(SubtopicClassroom(subtopic_id=sub_id, topic_id=top_id, name=st['name'], description=""))
                        all_subtopics.append((sub_id, st['name'], t['name'], cname, sname))
                db.commit()
            await asyncio.sleep(3)  # Respect 30 RPM
        
        logger.info(f"✅ Step 3 done: {len(all_topics)} topics, {len(all_subtopics)} subtopics created")
        
        # ── STEP 4: Generate Descriptions (batch 3 subtopics per call to save quota) ──
        batch_size = 3
        for i in range(0, len(all_subtopics), batch_size):
            batch = all_subtopics[i:i+batch_size]
            
            subtopic_list = ", ".join([f'"{st[1]} (Topic: {st[2]})"' for st in batch])
            prompt = f"""Subject: {batch[0][4]}, Chapter: {batch[0][3]}.
Generate detailed study descriptions for these subtopics: [{subtopic_list}].
For each subtopic, write: Intro, Body, and Conclusion.
Return JSON: {{"descriptions": [{{"subtopic_name": "...", "description": "Intro: ...\\n\\nBody: ...\\n\\nConclusion: ..."}}]}}"""
            
            res = await ask_llm(prompt)
            if res and 'descriptions' in res:
                for desc_item in res['descriptions']:
                    desc_name = desc_item.get('subtopic_name', '')
                    desc_text = desc_item.get('description', '')
                    # Match by name to the batch
                    for sub_id, sname, tname, cname, sjname in batch:
                        if sname.lower() in desc_name.lower() or desc_name.lower() in sname.lower():
                            sub_obj = db.query(SubtopicClassroom).filter(SubtopicClassroom.subtopic_id == sub_id).first()
                            if sub_obj and not sub_obj.description:
                                sub_obj.description = desc_text
                                break
                db.commit()
            await asyncio.sleep(3)  # Respect 30 RPM
        
        logger.info(f"✅ Step 4 done: descriptions generated")
        logger.info(f"🎉 Pipeline COMPLETED for '{paper_name}' — Total Groq calls: {call_count}")
        
    except Exception as e:
        logger.error(f"Pipeline error: {e}")
        db.rollback()
    finally:
        db.close()

# ── Bulk Auto-Generate Course Endpoint ────────────────────────────────────────

@router.post("/classroom/papers/{paper_id}/auto-generate", tags=["Classroom"])
async def auto_generate_course(paper_id: str, background_tasks: BackgroundTasks, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    paper = db.query(PaperClassroom).filter(PaperClassroom.paper_id == paper_id).first()
    if not paper:
        raise HTTPException(404, "Paper not found")
        
    exam = db.query(Exam).filter(Exam.exam_id == paper.exam_id).first()
    if not exam:
        raise HTTPException(404, "Exam not found")

    background_tasks.add_task(
        run_generation_pipeline,
        paper_id=paper_id,
        exam_id=exam.exam_id,
        exam_category=exam.category,
        exam_name=exam.name,
        paper_name=paper.name
    )
    
    return {"success": True, "message": "Generation started in the background. Please wait a few minutes and refresh."}


# ── Classroom Reel Generation Endpoint ───────────────────────────────────────

class TranscriptReq(BaseModel):
    language: Optional[str] = "English"

@router.post("/classroom/subtopics/{subtopic_id}/generate-transcript", tags=["Classroom"])
async def generate_subtopic_transcript(subtopic_id: str, req: TranscriptReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")

    topic = subtopic.topic
    chapter = topic.chapter if topic else None
    subject = chapter.subject if chapter else None

    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name if topic else subtopic.name
    lang = req.language or "English"

    raw = subtopic.description or subtopic.notes or subtopic.name
    import re as _re
    plain = _re.sub(r'<[^>]+>', '', raw)
    plain = _re.sub(r'\[IMAGE:[^\]]+\]', '', plain)
    plain = _re.sub(r'[#*`>_~]', '', plain).strip()[:2000]

    prompt = f"""You are an expert educational content creator.
Write a detailed, engaging 1-minute educational video narration script in {lang} for the following topic.

Subject: {subject_name} | Chapter: {chapter_name} | Topic: {topic_name}
Subtopic: {subtopic.name}

Source material:
{plain}

Requirements:
- Write in {lang} language only.
- CRITICAL: If the language is Hindi, write the dialogue strictly in Devanagari Unicode script (e.g. "भारत", "विज्ञान"). NEVER write in Hinglish (Hindi written using English/Latin alphabet, e.g. "Bharat", "vigyan"), as TTS engines pronounce Hinglish with a highly robotic/incorrect accent.
- CRITICAL: Spell out all numbers, place names, acronyms, and math symbols fully in spoken words of the target language (e.g. write "उन्नीस सौ सैंतालीस" instead of "1947", "प्रतिशत" / "percent" instead of "%", "किलोमीटर" instead of "km") so that ElevenLabs reads them with perfect professional pronunciation.
- Length: approximately 150-180 words (enough for 60 seconds of natural speech)
- Style: conversational, educational, engaging — like a knowledgeable teacher explaining to students
- Cover: introduction, key concepts, important facts, and a brief conclusion
- Use simple clear sentences, no bullet points, no scene labels — just flowing narration text
- Mix technical terms with simple explanations (like the reference style: use English terms with {lang} explanations)

Return ONLY the narration text, nothing else."""

    try:
        from app.services.llm import generate_simple_response
        transcript = await generate_simple_response(prompt, system_prompt=f"You are an expert educational narrator. Write natural, flowing narration in {lang}.")
        return {"success": True, "transcript": transcript.strip()}
    except Exception as e:
        logger.error(f"Error generating transcript: {e}")
        raise HTTPException(500, f"Failed to generate transcript: {str(e)}")


@router.post("/classroom/topics/{topic_id}/generate-transcript", tags=["Classroom"])
async def generate_topic_transcript(topic_id: str, req: TranscriptReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        TopicClassroom.topic_id == topic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")

    chapter = topic.chapter
    subject = chapter.subject if chapter else None

    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name
    lang = req.language or "English"

    raw = topic.script or topic.name
    import re as _re
    plain = _re.sub(r'<[^>]+>', '', raw)
    plain = _re.sub(r'\[IMAGE:[^\]]+\]', '', plain)
    plain = _re.sub(r'[#*`>_~]', '', plain).strip()[:2000]

    prompt = f"""You are an expert educational content creator.
Write a detailed, engaging 1-minute educational video narration script in {lang} for the following topic.

Subject: {subject_name} | Chapter: {chapter_name} | Topic: {topic_name}

Source material:
{plain}

Requirements:
- Write in {lang} language only.
- CRITICAL: If the language is Hindi, write the dialogue strictly in Devanagari Unicode script (e.g. "भारत", "विज्ञान"). NEVER write in Hinglish (Hindi written using English/Latin alphabet, e.g. "Bharat", "vigyan"), as TTS engines pronounce Hinglish with a highly robotic/incorrect accent.
- CRITICAL: Spell out all numbers, place names, acronyms, and math symbols fully in spoken words of the target language (e.g. write "उन्नीस सौ सैंतालीस" instead of "1947", "प्रतिशत" / "percent" instead of "%", "किलोमीटर" instead of "km") so that ElevenLabs reads them with perfect professional pronunciation.
- Length: approximately 150-180 words (enough for 60 seconds of natural speech)
- Style: conversational, educational, engaging — like a knowledgeable teacher explaining to students
- Cover: introduction, key concepts, important facts, and a brief conclusion
- Use simple clear sentences, no bullet points, no scene labels — just flowing narration text
- Mix technical terms with simple explanations (like the reference style: use English terms with {lang} explanations)

Return ONLY the narration text, nothing else."""

    try:
        from app.services.llm import generate_simple_response
        transcript = await generate_simple_response(prompt, system_prompt=f"You are an expert educational narrator. Write natural, flowing narration in {lang}.")
        return {"success": True, "transcript": transcript.strip()}
    except Exception as e:
        logger.error(f"Error generating topic transcript: {e}")
        raise HTTPException(500, f"Failed to generate transcript: {str(e)}")


@router.post("/classroom/pyqs/{pyq_set_id}/generate-transcript", tags=["Classroom"])
async def generate_pyq_transcript(pyq_set_id: str, req: TranscriptReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    from app.core.models import PYQSet, PYQQuestion
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ Set not found or access denied")

    questions = db.query(PYQQuestion).filter(PYQQuestion.pyq_set_id == pyq_set_id).order_by(PYQQuestion.created_at.asc()).all()
    if not questions:
        raise HTTPException(400, "No questions found in this PYQ set. Please upload a PDF/Excel first.")
        
    plain_qs = "\n".join(
        f"Question {i+1}: {q.question_text[:150]} -> Correct Answer: {q.correct_answer or 'See explanation'}"
        for i, q in enumerate(questions[:15])
    )
    
    lang = req.language or "English"

    prompt = f"""You are an expert educational content creator.
Write a detailed, engaging 3-minute educational video narration script in {lang} explaining the key concepts, questions, and answers from the following Previous Year Questions (PYQ) set.
Do NOT just list the questions. Write a flowing narrative lesson that explains the logic behind these questions and why the correct answers are what they are.

PYQ Set Name: {pyq_set.name}

Questions and Answers:
{plain_qs}

Requirements:
- Write in {lang} language only.
- CRITICAL: If the language is Hindi, write the dialogue strictly in Devanagari Unicode script (e.g. "भारत", "विज्ञान"). NEVER write in Hinglish (Hindi written using English/Latin alphabet, e.g. "Bharat", "vigyan"), as TTS engines pronounce Hinglish with a highly robotic/incorrect accent.
- CRITICAL: Spell out all numbers, place names, acronyms, and math symbols fully in spoken words of the target language (e.g. write "उन्नीस सौ सैंतालीस" instead of "1947", "प्रतिशत" / "percent" instead of "%", "किलोमीटर" instead of "km") so that ElevenLabs reads them with perfect professional pronunciation.
- Length: approximately 400-500 words (enough for a 3-minute natural spoken explanation)
- Style: conversational, educational, engaging — like a knowledgeable teacher explaining to students
- Cover: introduction, explanation of key question concepts, and a brief summary of what we learned
- Use simple clear sentences, no bullet points, no scene labels — just flowing narration text

Return ONLY the narration text, nothing else."""

    try:
        from app.services.llm import generate_simple_response
        transcript = await generate_simple_response(prompt, system_prompt=f"You are an expert educational narrator. Write natural, flowing narration in {lang}.")
        return {"success": True, "transcript": transcript.strip()}
    except Exception as e:
        logger.error(f"Error generating PYQ transcript: {e}")
        raise HTTPException(500, f"Failed to generate transcript: {str(e)}")


@router.post("/classroom/current-affairs/{ca_topic_id}/generate-transcript", tags=["Classroom"])
async def generate_ca_transcript(ca_topic_id: str, req: TranscriptReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    from app.core.models import CurrentAffairTopic
    ca_topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not ca_topic:
        raise HTTPException(404, "Current Affairs topic not found or access denied")

    raw = ca_topic.script or ca_topic.name
    import re as _re
    plain = _re.sub(r'<[^>]+>', '', raw)
    plain = _re.sub(r'\[IMAGE:[^\]]+\]', '', plain)
    plain = _re.sub(r'[#*`>_~]', '', plain).strip()[:2000]
    
    lang = req.language or "English"

    prompt = f"""You are an expert news and current affairs content creator.
Write a detailed, engaging 1-minute video narration script in {lang} explaining the following current affairs topic.
Do NOT just read the raw news text or facts. Tell an engaging story about this topic, highlighting why it is important, the context, and key details.

Topic Name: {ca_topic.name}

Source material / Facts:
{plain}

Requirements:
- Write in {lang} language only.
- CRITICAL: If the language is Hindi, write the dialogue strictly in Devanagari Unicode script (e.g. "भारत", "विज्ञान"). NEVER write in Hinglish (Hindi written using English/Latin alphabet, e.g. "Bharat", "vigyan"), as TTS engines pronounce Hinglish with a highly robotic/incorrect accent.
- CRITICAL: Spell out all numbers, place names, acronyms, and math symbols fully in spoken words of the target language (e.g. write "उन्नीस सौ सैंतालीस" instead of "1947", "प्रतिशत" / "percent" instead of "%", "किलोमीटर" instead of "km") so that ElevenLabs reads them with perfect professional pronunciation.
- Length: approximately 150-180 words (enough for 60 seconds of natural spoken news story)
- Style: professional, informative, engaging — like an expert journalist explaining to viewers
- Cover: hook/headline, key facts, context, and a brief takeaway
- Use simple clear sentences, no bullet points, no scene labels — just flowing narration text

Return ONLY the narration text, nothing else."""

    try:
        from app.services.llm import generate_simple_response
        transcript = await generate_simple_response(prompt, system_prompt=f"You are an expert news narrator. Write natural, flowing narration in {lang}.")
        return {"success": True, "transcript": transcript.strip()}
    except Exception as e:
        logger.error(f"Error generating CA transcript: {e}")
        raise HTTPException(500, f"Failed to generate transcript: {str(e)}")


class ClassroomReelReq(BaseModel):
    language: Optional[str] = "English"
    voice_id: Optional[str] = None
    topic_name: Optional[str] = ""
    transcript: Optional[str] = ""

@router.post("/classroom/subtopics/{subtopic_id}/generate-reel", tags=["Classroom"])
async def generate_classroom_reel(subtopic_id: str, req: ClassroomReelReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(
        SubtopicClassroom.subtopic_id == subtopic_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")

    topic = subtopic.topic
    chapter = topic.chapter if topic else None
    subject = chapter.subject if chapter else None
    paper = subject.paper if subject else None
    exam = paper.exam if paper else None

    exam_name = exam.name if exam else "General"
    subject_name = subject.name if subject else "General"
    chapter_name = chapter.name if chapter else "General"
    topic_name = topic.name if topic else subtopic.name
    lang = req.language or "English"

    # Prioritize requested transcript or database script over description/notes
    if req.transcript and req.transcript.strip():
        study_material = req.transcript.strip()
    else:
        study_material = subtopic.script or subtopic.description or subtopic.notes or subtopic.name

    # Generate structured scene-by-scene script
    from app.services.llm import generate_simple_response
    struct_prompt = f"""Convert the following study material into a structured scene-by-scene educational video reel script.

Subject: {subject_name}, Chapter: {chapter_name}, Topic: {topic_name}, Subtopic: {subtopic.name}
Exam: {exam_name}

STUDY MATERIAL:
{study_material[:3000]}

TARGET LANGUAGE FOR DIALOGUE: {lang}

🎙️ Dialogue Pronunciation Rules:
1. If the target language is Hindi, you MUST write the dialogue strictly in proper Devanagari Unicode script (e.g. "भारत", "प्रौद्योगिकी", "इतिहास"). NEVER write in Hinglish (Hindi written using English/Latin alphabet, e.g. "Bharat", "vigyan"), as TTS engines pronounce Hinglish with a highly robotic/incorrect accent.
2. CRITICAL: Spell out all numbers, place names, acronyms, and math symbols fully in spoken words of the target language (e.g. write "उन्नीस सौ सैंतालीस" instead of "1947", "प्रतिशत" / "percent" instead of "%", "किलोमीटर" instead of "km") so that ElevenLabs reads them with perfect professional pronunciation.

FORMAT FOR EACH SCENE (exactly this format):

🎬 Scene 1 (0-5 sec)
🎙️ Dialogue: [Narration in {lang}, 15-25 words, natural spoken sentence]
📸 Visuals / Footage: [Detailed English description for AI image generation]
🎥 Editing Notes: [Camera movement]

Each scene is 5 seconds.
Do NOT include any intro, outro, headers, or markdown wrappers. Only output the scene blocks."""

    structured_script = await generate_simple_response(struct_prompt, "You are a professional educational video script writer. Output only the scene blocks as requested.")

    from app.services.video_engine import assemble_advanced_reel
    import json
    from app.core.models import SocialContent

    res = await assemble_advanced_reel(
        structured_script,
        language=lang,
        voice_id=req.voice_id,
        bgm_style="cinematic"
    )

    video_url = res.get("video_url") if res else None
    scenes_data = res.get("scenes", []) if res else []

    if not video_url:
        raise HTTPException(500, "Reel generation failed")

    content_id = secrets.token_hex(8)

    # ── CLOUDFLARE R2 UPLOAD ──
    try:
        from app.services.r2_storage import upload_to_r2
        local_video_path = os.path.join(os.getcwd(), "uploads", "social", os.path.basename(video_url))
        
        # Key: reels/subtopic_{subtopic_id}/reel_{content_id}.mp4
        r2_key_unique = f"reels/subtopic_{subtopic_id}/reel_{content_id}.mp4"
        r2_key_latest = f"reels/subtopic_{subtopic_id}/latest.mp4"
        
        # Upload unique reel
        r2_url = upload_to_r2(local_video_path, r2_key_unique, "video/mp4")
        
        # Upload latest.mp4 for static referencing
        upload_to_r2(local_video_path, r2_key_latest, "video/mp4")
        
        if r2_url:
            logger.info(f"R2 Storage: Successfully saved classroom reel in Cloudflare R2! URL={r2_url}")
            video_url = r2_url
    except Exception as r2_err:
        logger.error(f"Failed to upload to Cloudflare R2 (falling back to local storage): {r2_err}")

    db_item = SocialContent(
        content_id=content_id,
        client_id=client["client_id"],
        content_type="reel",
        title=f"{subtopic.name} — {subject_name}",
        body=structured_script[:1000],
        media_url=video_url,
        scenes_json=json.dumps(scenes_data),
        metadata_json=json.dumps({
            "subtopic_id": subtopic_id,
            "exam_id": exam.exam_id if exam else "",
            "voice_id": req.voice_id,
            "language": lang,
            "script": structured_script
        })
    )
    db.add(db_item)
    db.commit()

    # Copy to subtopic-specific permanent directory for easy integration with other projects
    try:
        import shutil
        orig_filename = os.path.basename(res.get("video_url"))
        local_video_path = os.path.join(os.getcwd(), "uploads", "social", orig_filename)
        if os.path.exists(local_video_path):
            subtopic_reels_dir = os.path.join(os.getcwd(), "uploads", "reels", f"subtopic_{subtopic_id}")
            os.makedirs(subtopic_reels_dir, exist_ok=True)
            
            # Copy as a unique file
            dest_unique = os.path.join(subtopic_reels_dir, f"reel_{content_id}.mp4")
            shutil.copy2(local_video_path, dest_unique)
            
            # Copy as latest.mp4 for static integration
            dest_latest = os.path.join(subtopic_reels_dir, "latest.mp4")
            if os.path.exists(dest_latest):
                try: os.remove(dest_latest)
                except: pass
            shutil.copy2(local_video_path, dest_latest)
            
            logger.info(f"Classroom Reel physically saved to: {dest_unique} and {dest_latest}")
    except Exception as copy_err:
        logger.error(f"Failed to copy classroom reel to subtopic directory: {copy_err}")

    return {
        "success": True,
        "content_id": content_id,
        "video_url": video_url,
        "scenes": scenes_data,
        "script": structured_script
    }


# ── Granular Fetch Endpoints for Step-by-Step UI ─────────────────────────────

@router.get("/classroom/exams/{exam_id}/papers", tags=["Classroom"])
async def list_exam_papers(exam_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.exam_id == exam_id, Exam.client_id == client["client_id"]).first()
    if not exam:
        raise HTTPException(404, "Exam not found or access denied")
    papers = db.query(PaperClassroom).filter(PaperClassroom.exam_id == exam_id).order_by(PaperClassroom.created_at.desc()).all()
    return {"success": True, "papers": [p.to_dict() for p in papers]}


@router.get("/classroom/papers/{paper_id}/subjects", tags=["Classroom"])
async def list_paper_subjects(paper_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    paper = db.query(PaperClassroom).join(Exam).filter(PaperClassroom.paper_id == paper_id, Exam.client_id == client["client_id"]).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")
    subjects = db.query(Subject).filter(Subject.paper_id == paper_id).order_by(Subject.created_at.desc()).all()
    return {"success": True, "subjects": [s.to_dict() for s in subjects]}


@router.get("/classroom/subjects/{subject_id}/chapters", tags=["Classroom"])
async def list_subject_chapters(subject_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(Subject.subject_id == subject_id, Exam.client_id == client["client_id"]).first()
    if not subject:
        raise HTTPException(404, "Subject not found or access denied")
    chapters = db.query(ChapterClassroom).filter(ChapterClassroom.subject_id == subject_id).order_by(ChapterClassroom.created_at.asc()).all()
    return {"success": True, "chapters": [c.to_dict() for c in chapters]}


@router.get("/classroom/chapters/{chapter_id}/topics", tags=["Classroom"])
async def list_chapter_topics(chapter_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    chapter = db.query(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(ChapterClassroom.chapter_id == chapter_id, Exam.client_id == client["client_id"]).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found or access denied")
    topics = db.query(TopicClassroom).filter(TopicClassroom.chapter_id == chapter_id).order_by(TopicClassroom.created_at.asc()).all()
    return {"success": True, "topics": [t.to_dict() for t in topics]}


@router.get("/classroom/topics/{topic_id}/subtopics", tags=["Classroom"])
async def list_topic_subtopics(topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    topic = db.query(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(TopicClassroom.topic_id == topic_id, Exam.client_id == client["client_id"]).first()
    if not topic:
        raise HTTPException(404, "Topic not found or access denied")
    subtopics = db.query(SubtopicClassroom).filter(SubtopicClassroom.topic_id == topic_id).order_by(SubtopicClassroom.created_at.asc()).all()
    return {"success": True, "subtopics": [s.to_dict() for s in subtopics]}


@router.get("/classroom/subtopics/{subtopic_id}", tags=["Classroom"])
async def get_subtopic_details(subtopic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    subtopic = db.query(SubtopicClassroom).join(TopicClassroom).join(ChapterClassroom).join(Subject).join(PaperClassroom).join(Exam).filter(SubtopicClassroom.subtopic_id == subtopic_id, Exam.client_id == client["client_id"]).first()
    if not subtopic:
        raise HTTPException(404, "Subtopic not found or access denied")
        
    updated = False
    
    # Build context
    topic = db.query(TopicClassroom).filter(TopicClassroom.topic_id == subtopic.topic_id).first()
    chapter_name = ""
    subject_name = ""
    exam_name = ""
    topic_name = topic.name if topic else ""
    if topic:
        chapter = db.query(ChapterClassroom).filter(ChapterClassroom.chapter_id == topic.chapter_id).first()
        if chapter:
            chapter_name = chapter.name
            subject = db.query(Subject).filter(Subject.subject_id == chapter.subject_id).first()
            if subject:
                subject_name = subject.name
                if subject.exam_id:
                    exam = db.query(Exam).filter(Exam.exam_id == subject.exam_id).first()
                    if exam:
                        exam_name = exam.name
    context_str = f"{exam_name} {subject_name} {chapter_name} {topic_name}".strip()

    if not subtopic.image_url:
        subtopic.image_url = await generate_ai_image_and_upload(subtopic.name, subtitle="Subtopic", context=context_str)
        updated = True
    if not subtopic.banner_url:
        subtopic.banner_url = await generate_banner_image_and_upload(subtopic.name, context=context_str)
        updated = True
    if updated:
        db.commit()
        db.refresh(subtopic)
        
    return {"success": True, "subtopic": subtopic.to_dict()}


# ── AI Script Enhancer & Excel Syllabus Import Endpoints ─────────────────────

class EnhanceScriptReq(BaseModel):
    script: str
    language: str

@router.post("/classroom/enhance-script", tags=["Classroom"])
async def enhance_script(req: EnhanceScriptReq, client: dict = Depends(_require_client)):
    prompt = f"""You are an elite educational content optimizer, dramatic scriptwriter, and ElevenLabs custom voice narrator.
Your goal is to enhance the following raw script to make it sound incredibly human, highly professional, deeply expressive, and captivating for a 60-second social media reel.

Requirements:
1. Hook: Make the opening sentence extremely catchy and educational.
2. Expression & Tone Markers:
   - CRITICAL: Aggressively and naturally inject emotional/pacing voice cues inside square brackets to guide ElevenLabs.
   - Examples of voice cues: `[thoughtful]`, `[excited]`, `[clears throat]`, `[dramatic pause]`, `[sighs]`, `[whispering]`, `[confident]`, `[laughs]`.
   - Add these cues before key segments to direct the tone (e.g., "[thoughtful] Did you know that...", "[excited] This is where the magic happens!", "[dramatic pause] But here is the catch...").
3. Flow: Improve sentence structure, pacing, and vocabulary for clear narration.
4. Spelling/TTS Formatting:
   - CRITICAL: Spell out all numbers, acronyms, math symbols, and percentages in spoken words of the target language (e.g., write "percent" / "प्रतिशत" instead of "%", "उन्नीस सौ सैंतालीस" instead of "1947").
   - CRITICAL: If the target language is Hindi, use strictly proper Devanagari Unicode script. Do NOT use Hinglish!
5. Length: Keep the length around 140 to 180 words, perfect for a 60-second voiceover.
6. Format: Return ONLY the enhanced narration script containing the text and the bracketed voice cues. Do not add any bullet points, labels like "Scene 1", intro, or outro.

Target Language: {req.language}
Original Script:
{req.script}

Enhanced Script:"""
    try:
        from app.services.llm import generate_simple_response
        enhanced_text = await generate_simple_response(prompt, system_prompt="You are a professional voiceover script enhancer.")
        return {"success": True, "enhanced_script": enhanced_text.strip()}
    except Exception as e:
        logger.error(f"Error enhancing script: {e}")
        raise HTTPException(500, f"Failed to enhance script: {str(e)}")


def parse_video_length(val) -> Optional[int]:
    if val is None:
        return None
    val_str = str(val).strip().lower()
    if not val_str:
        return None
    
    import re
    # Extract numeric part
    match = re.search(r"(\d+(?:\.\d+)?)", val_str)
    if not match:
        return None
    num = float(match.group(1))
    
    # Detect minutes
    # E.g. "1 min", "2 minutes", "1.5m", "1.5 min", "1.5 mins"
    if "min" in val_str or re.search(r"\b(m|mins|minute|minutes)\b", val_str) or (val_str.endswith("m") and not val_str.endswith("am") and not val_str.endswith("pm")):
        return int(num * 60)
    
    return int(num)


@router.post("/classroom/subjects/{subject_id}/upload-index", tags=["Classroom"])
async def upload_subject_index(
    subject_id: str,
    file: UploadFile = File(...),
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    # 1. Fetch and validate Subject ownership
    subject = db.query(Subject).join(PaperClassroom).join(Exam).filter(
        Subject.subject_id == subject_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not subject:
        raise HTTPException(404, "Subject not found or access denied")

    # 2. Read and parse file bytes using openpyxl
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise HTTPException(400, f"Invalid Excel file: {str(e)}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "The Excel file is empty")

    start_row = 0
    first_row = rows[0]
    
    # Column indices mapping (default values assume: 0=Chapter, 1=Topic, 2=Subtopic, 3=Script)
    chapter_col = 0
    topic_col = 1
    subtopic_col = 2
    script_col = 3
    length_col = -1

    # Check if first row contains column headers
    if first_row and any(isinstance(val, str) and any(h in val.lower() for h in ["chapter", "topic", "subtopic", "subject", "script", "length", "duration"]) for val in first_row):
        start_row = 1
        
        # Reset all indices to -1 when header row is detected to map strictly based on headers
        chapter_col = -1
        topic_col = -1
        subtopic_col = -1
        script_col = -1
        length_col = -1

        # Identify column mappings from header
        for idx, val in enumerate(first_row):
            if val is None or not isinstance(val, str):
                continue
            val_lower = val.strip().lower()
            if "subtopic" in val_lower:
                subtopic_col = idx
            elif "topic" in val_lower:
                topic_col = idx
            elif "chapter" in val_lower:
                chapter_col = idx
            elif "script" in val_lower:
                script_col = idx
            elif "length" in val_lower or "duration" in val_lower:
                length_col = idx
        
        # Set default values for required columns if they were not explicitly mapped
        if chapter_col == -1:
            chapter_col = 0
        if topic_col == -1:
            topic_col = 1
        if subtopic_col == -1:
            subtopic_col = 2
        if script_col == -1:
            script_col = 3


    chapters_created = 0
    topics_created = 0
    subtopics_created = 0

    # Caches to look up existing items
    chapter_cache = {}
    topic_cache = {}

    existing_chapters = db.query(ChapterClassroom).filter(ChapterClassroom.subject_id == subject_id).all()
    for ch in existing_chapters:
        chapter_cache[ch.name.strip().lower()] = ch.chapter_id
        for tp in ch.topics:
            topic_cache[(ch.chapter_id, tp.name.strip().lower())] = tp.topic_id

    for r_idx in range(start_row, len(rows)):
        row = rows[r_idx]
        if not row or all(v is None for v in row):
            continue

        ch_name = ""
        if chapter_col != -1 and chapter_col < len(row) and row[chapter_col] is not None:
            ch_name = str(row[chapter_col]).strip()

        tp_name = ""
        if topic_col != -1 and topic_col < len(row) and row[topic_col] is not None:
            tp_name = str(row[topic_col]).strip()

        st_name = ""
        if subtopic_col != -1 and subtopic_col < len(row) and row[subtopic_col] is not None:
            st_name = str(row[subtopic_col]).strip()

        script_text = None
        if script_col != -1 and script_col < len(row) and row[script_col] is not None:
            script_text = str(row[script_col]).strip()

        video_length_val = None
        if length_col != -1 and length_col < len(row) and row[length_col] is not None:
            video_length_val = parse_video_length(row[length_col])

        if not ch_name:
            continue

        ch_key = ch_name.lower()
        if ch_key not in chapter_cache:
            chapter_id = "chapter-" + secrets.token_hex(8)
            new_ch = ChapterClassroom(
                chapter_id=chapter_id,
                subject_id=subject_id,
                name=ch_name,
                created_at=datetime.utcnow()
            )
            db.add(new_ch)
            db.commit()
            chapter_cache[ch_key] = chapter_id
            chapters_created += 1

        chapter_id = chapter_cache[ch_key]

        if not tp_name:
            continue

        tp_key = (chapter_id, tp_name.lower())
        if tp_key not in topic_cache:
            topic_id = "topic-" + secrets.token_hex(8)
            new_tp = TopicClassroom(
                topic_id=topic_id,
                chapter_id=chapter_id,
                name=tp_name,
                created_at=datetime.utcnow()
            )
            # Add topic script and length if we don't have subtopics on this row,
            # or use them as initial fallback if they are provided.
            new_tp.script = script_text
            new_tp.video_length = video_length_val
            db.add(new_tp)
            db.commit()
            topic_cache[tp_key] = topic_id
            topics_created += 1
        else:
            # Topic already exists.
            topic_id = topic_cache[tp_key]
            topic_obj = db.query(TopicClassroom).filter(TopicClassroom.topic_id == topic_id).first()
            if topic_obj:
                # If there is no subtopic name on this row, update topic script and video_length (explicit overwrite)
                if not st_name:
                    if script_text:
                        topic_obj.script = script_text
                    if video_length_val is not None:
                        topic_obj.video_length = video_length_val
                    db.commit()
                # Otherwise, if subtopic is present, only write script/video_length as a fallback if topic currently has none
                else:
                    updated = False
                    if not topic_obj.script and script_text:
                        topic_obj.script = script_text
                        updated = True
                    if topic_obj.video_length is None and video_length_val is not None:
                        topic_obj.video_length = video_length_val
                        updated = True
                    if updated:
                        db.commit()

        topic_id = topic_cache[tp_key]

        if not st_name:
            continue

        existing_st = db.query(SubtopicClassroom).filter(
            SubtopicClassroom.topic_id == topic_id,
            SubtopicClassroom.name == st_name
        ).first()

        if not existing_st:
            subtopic_id = "subtopic-" + secrets.token_hex(8)
            new_st = SubtopicClassroom(
                subtopic_id=subtopic_id,
                topic_id=topic_id,
                name=st_name,
                description="",
                notes="",
                script=script_text,
                created_at=datetime.utcnow()
            )
            db.add(new_st)
            db.commit()
            subtopics_created += 1
        else:
            if script_text:
                existing_st.script = script_text
                db.commit()

    return {
        "success": True,
        "chapters_created": chapters_created,
        "topics_created": topics_created,
        "subtopics_created": subtopics_created
    }


# ══════════════════════════════════════════════════════════════════════════════
# ── CURRENT AFFAIRS SECTION ────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

class CreateCATopicReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    script: Optional[str] = None
    image_url: Optional[str] = ""

class UpdateCATopicReq(BaseModel):
    name: Optional[str] = None
    script: Optional[str] = None
    image_url: Optional[str] = None


@router.get("/classroom/current-affairs", tags=["Classroom CA"])
async def list_ca_topics(client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """List all Current Affairs topics for this client."""
    topics = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.client_id == client["client_id"]
    ).order_by(CurrentAffairTopic.created_at.desc()).all()
    return {"success": True, "topics": [t.to_dict() for t in topics]}


@router.post("/classroom/current-affairs", tags=["Classroom CA"])
async def create_ca_topic(req: CreateCATopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Create a new Current Affairs topic."""
    image_url = req.image_url
    if not image_url:
        image_url = generate_premium_image_locally(req.name, subtitle="Current Affair")
    elif image_url.startswith("http") and not image_url.startswith("/uploads"):
        image_url = download_and_save_image(image_url, fallback_keyword=req.name)

    ca_topic_id = "ca-" + secrets.token_hex(8)
    topic = CurrentAffairTopic(
        ca_topic_id=ca_topic_id,
        client_id=client["client_id"],
        name=req.name,
        script=req.script,
        image_url=image_url,
        created_at=datetime.utcnow()
    )
    db.add(topic)
    db.commit()
    db.refresh(topic)
    return {"success": True, "topic": topic.to_dict()}


@router.put("/classroom/current-affairs/{ca_topic_id}", tags=["Classroom CA"])
async def update_ca_topic(ca_topic_id: str, req: UpdateCATopicReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Update a Current Affairs topic name or script."""
    topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
    if req.name is not None:
        topic.name = req.name
        # Automatically update/regenerate the cover image to match the new name
        topic.image_url = generate_premium_image_locally(req.name, subtitle="Current Affair")
    if req.script is not None:
        topic.script = req.script
    if req.image_url is not None:
        image_url = req.image_url
        if image_url and image_url.startswith("http") and not image_url.startswith("/uploads"):
            image_url = download_and_save_image(image_url, fallback_keyword=topic.name)
        topic.image_url = image_url
        
    db.commit()
    db.refresh(topic)
    return {"success": True, "topic": topic.to_dict()}


@router.delete("/classroom/current-affairs/{ca_topic_id}", tags=["Classroom CA"])
async def delete_ca_topic(ca_topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Delete a Current Affairs topic and all its reels."""
    topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
    db.delete(topic)
    db.commit()
    return {"success": True, "message": "Topic deleted"}


@router.post("/classroom/current-affairs/{ca_topic_id}/upload-pdf", tags=["Classroom CA"])
async def upload_ca_pdf(
    ca_topic_id: str,
    file: UploadFile = File(...),
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    """Upload a PDF/Excel file for a Current Affairs topic. Extracts text to use as script material."""
    topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")

    filename = file.filename or "upload.pdf"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    upload_dir = os.path.join(os.getcwd(), "uploads", "ca_pdfs")
    os.makedirs(upload_dir, exist_ok=True)

    safe_name = f"{ca_topic_id}_{secrets.token_hex(4)}.{ext}"
    save_path = os.path.join(upload_dir, safe_name)

    contents = await file.read()
    with open(save_path, "wb") as f_out:
        f_out.write(contents)

    # Extract text from the file to use as script material
    extracted_text = ""
    try:
        if ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            rows_text = []
            for row in sheet.iter_rows(values_only=True):
                row_str = " | ".join(str(v) for v in row if v is not None)
                if row_str.strip():
                    rows_text.append(row_str)
            extracted_text = "\n".join(rows_text[:200])  # Cap at 200 rows
        elif ext == "pdf":
            extracted_text = extract_text_from_pdf(contents, max_pages=20)
            if not extracted_text.strip():
                logger.info("CA PDF text extraction returned empty text. Falling back to scanned PDF text OCR...")
                extracted_text = await extract_text_from_scanned_pdf(contents, max_pages=15)
    except Exception as parse_err:
        logger.warning(f"CA PDF text extraction failed: {parse_err}")

    # Save file info and update script if no script already set
    topic.pdf_filename = filename
    topic.pdf_path = save_path
    if not topic.script and extracted_text:
        topic.script = extracted_text[:5000]  # Cap script at 5000 chars
    db.commit()
    db.refresh(topic)
    return {
        "success": True,
        "message": "PDF uploaded successfully",
        "extracted_chars": len(extracted_text),
        "topic": topic.to_dict()
    }


@router.post("/classroom/current-affairs/{ca_topic_id}/generate-reel", tags=["Classroom CA"])
async def generate_ca_reel(
    ca_topic_id: str,
    language: str = Form("English"),
    voice_id: Optional[str] = Form(None),
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    """Generate a short educational reel for a Current Affairs topic."""
    topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")

    source_material = topic.script or topic.name
    lang = language or "English"

    from app.services.llm import generate_simple_response
    struct_prompt = f"""Convert the following Current Affairs material into a structured scene-by-scene educational video reel script.

Topic: {topic.name}

MATERIAL:
{source_material[:3000]}

TARGET LANGUAGE FOR DIALOGUE: {lang}

🎙️ Dialogue Pronunciation Rules:
1. If the target language is Hindi, write the dialogue strictly in proper Devanagari Unicode script. NEVER use Hinglish.
2. Spell out all numbers, acronyms, and math symbols fully in spoken words of the target language.

FORMAT FOR EACH SCENE (exactly this format):

🎬 Scene 1 (0-5 sec)
🎙️ Dialogue: [Narration in {lang}, 15-25 words, natural spoken sentence]
📸 Visuals / Footage: [Detailed English description for AI image generation]
🎥 Editing Notes: [Camera movement]

Each scene is 5 seconds. Generate 6-10 scenes for a ~30-60 second reel.
Do NOT include any intro, outro, headers, or markdown wrappers. Only output the scene blocks."""

    structured_script = await generate_simple_response(struct_prompt, "You are a professional educational video script writer. Output only the scene blocks as requested.")

    from app.services.video_engine import assemble_advanced_reel
    import json
    res = await assemble_advanced_reel(
        structured_script,
        language=lang,
        voice_id=voice_id,
        bgm_style="cinematic"
    )

    video_url = res.get("video_url") if res else None
    if not video_url:
        raise HTTPException(500, "Reel generation failed")

    reel_id = secrets.token_hex(8)

    # Attempt R2 upload
    try:
        from app.services.r2_storage import upload_to_r2
        local_video_path = os.path.join(os.getcwd(), "uploads", "social", os.path.basename(video_url))
        r2_key = f"reels/ca_{ca_topic_id}/reel_{reel_id}.mp4"
        r2_url = upload_to_r2(local_video_path, r2_key, "video/mp4")
        if r2_url:
            video_url = r2_url
    except Exception as r2_err:
        logger.error(f"CA reel R2 upload failed (using local): {r2_err}")

    reel = CurrentAffairReel(
        reel_id=reel_id,
        ca_topic_id=ca_topic_id,
        client_id=client["client_id"],
        media_url=video_url,
        script=structured_script[:2000],
        created_at=datetime.utcnow()
    )
    db.add(reel)
    db.commit()
    db.refresh(reel)

    return {
        "success": True,
        "reel_id": reel_id,
        "video_url": video_url,
        "script": structured_script
    }


@router.get("/classroom/current-affairs/{ca_topic_id}/reels", tags=["Classroom CA"])
async def list_ca_reels(ca_topic_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """List all reels generated for a Current Affairs topic."""
    topic = db.query(CurrentAffairTopic).filter(
        CurrentAffairTopic.ca_topic_id == ca_topic_id,
        CurrentAffairTopic.client_id == client["client_id"]
    ).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
    reels = db.query(CurrentAffairReel).filter(
        CurrentAffairReel.ca_topic_id == ca_topic_id
    ).order_by(CurrentAffairReel.created_at.desc()).all()
    return {"success": True, "reels": [r.to_dict() for r in reels]}


@router.delete("/classroom/current-affairs/reels/{reel_id}", tags=["Classroom CA"])
async def delete_ca_reel(reel_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Delete a specific Current Affairs reel."""
    reel = db.query(CurrentAffairReel).filter(
        CurrentAffairReel.reel_id == reel_id,
        CurrentAffairReel.client_id == client["client_id"]
    ).first()
    if not reel:
        raise HTTPException(404, "Reel not found")
    db.delete(reel)
    db.commit()
    return {"success": True, "message": "Reel deleted"}


# ══════════════════════════════════════════════════════════════════════════════
# ── PYQ (PREVIOUS YEAR QUESTIONS) SECTION ─────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

class CreatePYQSetReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)


@router.get("/classroom/pyq-sets", tags=["Classroom PYQ"])
async def list_pyq_sets(client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """List all PYQ sets for this client."""
    sets = db.query(PYQSet).filter(
        PYQSet.client_id == client["client_id"]
    ).order_by(PYQSet.created_at.desc()).all()
    return {"success": True, "pyq_sets": [s.to_dict() for s in sets]}


@router.post("/classroom/pyq-sets", tags=["Classroom PYQ"])
async def create_pyq_set(req: CreatePYQSetReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Create a new named PYQ set, e.g. 'BPSC 72'."""
    pyq_set_id = "pyq-" + secrets.token_hex(8)
    pyq_set = PYQSet(
        pyq_set_id=pyq_set_id,
        client_id=client["client_id"],
        name=req.name,
        created_at=datetime.utcnow()
    )
    db.add(pyq_set)
    db.commit()
    db.refresh(pyq_set)
    return {"success": True, "pyq_set": pyq_set.to_dict()}


@router.delete("/classroom/pyq-sets/{pyq_set_id}", tags=["Classroom PYQ"])
async def delete_pyq_set(pyq_set_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Delete a PYQ set and all its questions."""
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")
    db.delete(pyq_set)
    db.commit()
    return {"success": True, "message": "PYQ set deleted"}


class UpdatePYQSetReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)


@router.put("/classroom/pyq-sets/{pyq_set_id}", tags=["Classroom PYQ"])
async def update_pyq_set(pyq_set_id: str, req: UpdatePYQSetReq, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Update a PYQ set's name."""
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")
    pyq_set.name = req.name
    db.commit()
    db.refresh(pyq_set)
    return {"success": True, "pyq_set": pyq_set.to_dict()}


@router.post("/classroom/pyq-sets/{pyq_set_id}/reset", tags=["Classroom PYQ"])
async def reset_pyq_set(pyq_set_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """
    Delete any uploaded PDF/Excel files for this set and clear all its questions from database.
    """
    import glob
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")
        
    # 1. Delete files on disk matching the pattern
    pattern = os.path.join(os.getcwd(), "uploads", "pyq_files", f"{pyq_set_id}_*")
    for f in glob.glob(pattern):
        try:
            if os.path.exists(f):
                os.remove(f)
        except Exception as fe:
            logger.warning(f"Failed to remove file {f}: {fe}")
            
    # 2. Delete all questions from DB
    db.query(PYQQuestion).filter(PYQQuestion.pyq_set_id == pyq_set_id).delete()
    
    # 3. Reset overview status
    pyq_set.overview_generated = False
    
    db.commit()
    return {"success": True, "message": "PYQ set reset successfully. All questions and uploaded files have been deleted."}


def parse_mcqs_rule_based(text: str) -> list:
    import re
    # Split on question numbers like "1. ", "2. ", "150. " followed by letters/questions/newlines
    pattern = r'\n\s*(\d{1,3})\.\s*(?=\n|[A-Z]|\?)'
    parts = re.split(pattern, text)
    
    questions = []
    i = 1
    while i < len(parts) - 1:
        q_num = parts[i]
        q_body = parts[i+1].strip()
        i += 2
        
        # Check if there is Hindi text (Devanagari range: \u0900-\u097F)
        hindi_match = re.search(r'[\u0900-\u097f]', q_body)
        
        english_part = q_body
        hindi_part = ""
        
        if hindi_match:
            split_idx = hindi_match.start()
            before_hindi = q_body[:split_idx]
            after_hindi = q_body[split_idx:]
            
            # Find the last newline in before_hindi to separate cleanly
            last_nl = before_hindi.rfind('\n')
            if last_nl != -1:
                english_part = before_hindi[:last_nl].strip()
                hindi_part = (before_hindi[last_nl:] + after_hindi).strip()
            else:
                english_part = before_hindi.strip()
                hindi_part = after_hindi.strip()
        
        # Helper to extract options
        def extract_opts(part_text):
            opt_pattern = r'\(([A-E])\)\s*(.*?)(?=\s*\([A-E]\)|\Z)'
            matches = re.findall(opt_pattern, part_text, re.DOTALL)
            
            opts_dict = {}
            for letter, o_text in matches:
                o_clean = re.sub(r'\s+', ' ', o_text).strip()
                if o_clean:
                    opts_dict[letter.upper()] = o_clean
            
            first_opt = re.search(r'\([A-E]\)', part_text)
            q_text = part_text
            if first_opt:
                q_text = part_text[:first_opt.start()].strip()
            q_text = re.sub(r'\s+', ' ', q_text).strip()
            
            return q_text, opts_dict

        eng_q, eng_opts = extract_opts(english_part)
        hin_q, hin_opts = extract_opts(hindi_part) if hindi_part else ("", {})
        
        # Combine bilingual text
        combined_q = eng_q
        if hin_q:
            combined_q += " / " + hin_q
            
        combined_opts = []
        for letter in ['A', 'B', 'C', 'D', 'E']:
            opts_for_letter = []
            if letter in eng_opts:
                opts_for_letter.append(eng_opts[letter])
            if letter in hin_opts:
                opts_for_letter.append(hin_opts[letter])
                
            if opts_for_letter:
                combined_opts.append(" / ".join(opts_for_letter))
                
        if len(combined_opts) >= 2:
            questions.append({
                "question": combined_q,
                "options": combined_opts,
                "correct": ""
            })
            
    return questions


@router.post("/classroom/pyq-sets/{pyq_set_id}/upload-pdf", tags=["Classroom PYQ"])
async def upload_pyq_pdf(
    pyq_set_id: str,
    file: UploadFile = File(...),
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    """
    Upload a PDF/Excel with PYQ questions.
    Excel format expected: columns → Question | Option A | Option B | Option C | Option D | Correct Answer
    PDF: Text is extracted and AI is used to parse MCQ questions.
    """
    import json as _json
    import re as _re

    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")

    filename = file.filename or "upload"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    contents = await file.read()

    # Save uploaded file to disk
    upload_dir = os.path.join(os.getcwd(), "uploads", "pyq_files")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"{pyq_set_id}_{secrets.token_hex(4)}.{ext}"
    save_path = os.path.join(upload_dir, safe_name)
    try:
        with open(save_path, "wb") as f_out:
            f_out.write(contents)
    except Exception as save_err:
        logger.warning(f"Failed to save PYQ file to disk: {save_err}")

    parsed_questions = []

    if ext in ("xlsx", "xls"):
        # ── Excel parsing ──
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(400, "Empty Excel file")

            # Detect header row
            start_row = 0
            q_col, oa_col, ob_col, oc_col, od_col, ans_col = 0, 1, 2, 3, 4, 5
            first_row = rows[0]
            if first_row and any(
                isinstance(v, str) and any(h in str(v).lower() for h in ["question", "option", "answer", "correct"])
                for v in first_row if v
            ):
                start_row = 1
                for idx, v in enumerate(first_row):
                    if v is None: continue
                    vl = str(v).strip().lower()
                    if "question" in vl: q_col = idx
                    elif "option a" in vl or vl == "a": oa_col = idx
                    elif "option b" in vl or vl == "b": ob_col = idx
                    elif "option c" in vl or vl == "c": oc_col = idx
                    elif "option d" in vl or vl == "d": od_col = idx
                    elif "answer" in vl or "correct" in vl: ans_col = idx

            for row in rows[start_row:]:
                if not row or all(v is None for v in row):
                    continue
                def get_cell(col):
                    if col < len(row) and row[col] is not None:
                        return str(row[col]).strip()
                    return ""
                q_text = get_cell(q_col)
                if not q_text:
                    continue
                options = [o for o in [get_cell(oa_col), get_cell(ob_col), get_cell(oc_col), get_cell(od_col)] if o]
                correct = get_cell(ans_col)
                parsed_questions.append({"question": q_text, "options": options, "correct": correct})
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Excel parse error: {str(e)}")

    elif ext == "pdf":
        # Extract text from digital PDF
        raw_text = extract_text_from_pdf(contents, max_pages=35)
        if not raw_text.strip():
            logger.info("PYQ PDF text extraction returned empty text. Falling back to scanned PDF OCR MCQ extraction...")
            try:
                parsed_questions = await extract_mcqs_from_scanned_pdf(contents, max_pages=25)
            except Exception as ocr_err:
                logger.error(f"Multimodal OCR MCQ extraction failed: {ocr_err}")
                parsed_questions = []
            
            if not parsed_questions:
                raise HTTPException(
                    400,
                    "No text could be extracted from this PDF, and multimodal OCR extraction failed. Please check if the PDF contains clear pages with readable questions."
                )
        else:
            # 1. Try rule-based bilingual parser first (offline, instant, bypasses API limits)
            try:
                parsed_questions = parse_mcqs_rule_based(raw_text)
                logger.info(f"Rule-based parser extracted {len(parsed_questions)} questions from PDF.")
            except Exception as rule_err:
                logger.error(f"Rule-based parsing failed: {rule_err}")
                parsed_questions = []

            # 2. If rule-based parser yields very few questions (< 3), fall back to chunk-based AI parsing
            if len(parsed_questions) < 3:
                logger.info("Rule-based parser extracted < 3 questions. Falling back to AI chunk-based parsing...")
                parsed_questions = []
                chunk_size = 10000
                overlap = 1500
                seen_questions = set()
                start = 0

                from app.services.llm import generate_simple_response

                while start < len(raw_text):
                    end = min(start + chunk_size, len(raw_text))
                    chunk_text = raw_text[start:end]

                    if end < len(raw_text):
                        last_space = chunk_text.rfind(' ')
                        if last_space != -1 and last_space > chunk_size - 1000:
                            chunk_text = chunk_text[:last_space]
                            end = start + last_space

                    if len(chunk_text.strip()) > 100:
                        try:
                            parse_prompt = f"""Extract all MCQ (Multiple Choice Questions) from the following exam text chunk.
Return a strict JSON array in this format:
[
  {{
    "question": "Question text here",
    "options": ["Option A", "Option B", "Option C", "Option D"],
    "correct": "Correct option text"
  }}
]
If correct answer is not shown, leave "correct" as empty string.
Do NOT include markdown wrappers, only return the JSON array.

EXAM TEXT CHUNK:
{chunk_text}"""

                            # Use OpenAI (gpt-4o-mini) as fallback if Groq (the default LLM provider) is rate limited/fails
                            try:
                                raw_response = await generate_simple_response(parse_prompt, "You are an MCQ extraction expert. Return only valid JSON.")
                            except Exception as default_llm_err:
                                logger.warning(f"Default LLM provider failed: {default_llm_err}. Trying OpenAI fallback...")
                                from openai import AsyncOpenAI
                                from app.core.config import settings
                                openai_client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
                                resp = await openai_client.chat.completions.create(
                                    model="gpt-4o-mini",
                                    messages=[
                                        {"role": "system", "content": "You are an MCQ extraction expert. Return only valid JSON."},
                                        {"role": "user", "content": parse_prompt}
                                    ],
                                    max_tokens=2048,
                                    temperature=0.1
                                )
                                raw_response = resp.choices[0].message.content.strip()

                            clean = _re.sub(r'```json\s*', '', raw_response)
                            clean = _re.sub(r'\s*```', '', clean).strip()

                            # Extract array bracket contents
                            first_bracket = clean.find('[')
                            last_bracket = clean.rfind(']')
                            if first_bracket != -1 and last_bracket != -1:
                                clean = clean[first_bracket:last_bracket+1]

                            chunk_questions = _json.loads(clean)
                            if isinstance(chunk_questions, list):
                                for q in chunk_questions:
                                    q_text = q.get("question", "").strip()
                                    if q_text:
                                        norm = _re.sub(r'\s+', ' ', q_text).lower()
                                        if norm not in seen_questions:
                                            seen_questions.add(norm)
                                            parsed_questions.append(q)
                        except Exception as parse_err:
                            logger.error(f"AI MCQ parsing failed for chunk starting at {start}: {parse_err}")

                    start = end - overlap
                    if start >= len(raw_text) or end >= len(raw_text):
                        break
    else:
        raise HTTPException(400, "Only PDF or Excel (.xlsx/.xls) files are supported.")

    if not parsed_questions:
        return {"success": True, "questions_added": 0, "message": "No questions found in file."}

    added = 0
    for q in parsed_questions:
        q_text = q.get("question", "").strip()
        if not q_text:
            continue
        options = q.get("options", [])
        correct = q.get("correct", "")
        q_id = "pyqq-" + secrets.token_hex(8)
        new_q = PYQQuestion(
            question_id=q_id,
            pyq_set_id=pyq_set_id,
            question_text=q_text,
            options_json=_json.dumps(options),
            correct_answer=correct,
            pdf_filename=filename,
            created_at=datetime.utcnow()
        )
        db.add(new_q)
        added += 1
    db.commit()

    return {"success": True, "questions_added": added, "message": f"{added} questions imported from {filename}"}


@router.get("/classroom/pyq-sets/{pyq_set_id}/questions", tags=["Classroom PYQ"])
async def list_pyq_questions(pyq_set_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """List all questions in a PYQ set."""
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")
    questions = db.query(PYQQuestion).filter(
        PYQQuestion.pyq_set_id == pyq_set_id
    ).order_by(PYQQuestion.created_at.asc()).all()
    return {"success": True, "questions": [q.to_dict() for q in questions]}


@router.delete("/classroom/pyq-sets/questions/{question_id}", tags=["Classroom PYQ"])
async def delete_pyq_question(question_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Delete a single PYQ question."""
    q = db.query(PYQQuestion).join(PYQSet).filter(
        PYQQuestion.question_id == question_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not q:
        raise HTTPException(404, "Question not found")
    db.delete(q)
    db.commit()
    return {"success": True, "message": "Question deleted"}


@router.post("/classroom/pyq-sets/{pyq_set_id}/generate-overview", tags=["Classroom PYQ"])
async def generate_pyq_overview(pyq_set_id: str, language: str = "English", client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """
    AI generates full solutions and explanations for all questions in a PYQ set.
    Updates each question's explanation field in the database.
    """
    import json as _json

    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")

    questions = db.query(PYQQuestion).filter(
        PYQQuestion.pyq_set_id == pyq_set_id
    ).order_by(PYQQuestion.created_at.asc()).all()

    if not questions:
        raise HTTPException(400, "No questions found in this PYQ set. Please upload a PDF/Excel first.")

    # Process question-by-question as requested
    total_explained = 0

    for idx, q in enumerate(questions):
        # Skip if already explained to support resuming from rate limits
        if q.explanation and q.explanation.strip():
            total_explained += 1
            continue

        prompt = f"""You are an expert exam tutor for competitive exams like BPSC, UPSC, SSC.
For the question below, confirm/determine the correct answer and provide a detailed explanation in 3-5 sentences.
CRITICAL: Write the explanation strictly in the {language} language.

Question: {q.question_text}
Options: {', '.join(q.options) if q.options else 'N/A'}
Reported Correct: {q.correct_answer or 'Unknown'}

Return a strict JSON object (NOT a list/array):
{{
  "correct_answer": "Confirm or write correct answer option here",
  "explanation": "Detailed explanation here..."
}}

Return ONLY the JSON object. Do NOT wrap in markdown code blocks."""

        success = False
        res = {}

        try:
            # Explicitly call Groq as requested
            raw = await generate_groq_response(prompt, "You are an expert exam solutions generator. Return only valid JSON.")
            import re as _re
            clean = _re.sub(r'```json\s*', '', raw)
            clean = _re.sub(r'\s*```', '', clean).strip()
            first_brace = clean.find('{')
            last_brace = clean.rfind('}')
            if first_brace != -1 and last_brace != -1:
                clean = clean[first_brace:last_brace+1]
            res = _json.loads(clean)
            success = True
        except Exception as groq_err:
            logger.warning(f"Groq overview generation failed for Q{idx+1} ({q.question_id}): {groq_err}. Trying Gemini fallback...")
            try:
                # Fall back to Gemini 2.5 Flash
                from app.core.config import settings
                import httpx
                url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={settings.GEMINI_API_KEY}"
                payload = {
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 1024}
                }
                async with httpx.AsyncClient(timeout=30.0) as hc:
                    r = await hc.post(url, json=payload)
                    if r.status_code == 200:
                        raw = r.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
                        import re as _re
                        clean = _re.sub(r'```json\s*', '', raw)
                        clean = _re.sub(r'\s*```', '', clean).strip()
                        first_brace = clean.find('{')
                        last_brace = clean.rfind('}')
                        if first_brace != -1 and last_brace != -1:
                            clean = clean[first_brace:last_brace+1]
                        res = _json.loads(clean)
                        success = True
                        logger.info(f"Successfully generated solution for Q{idx+1} using Gemini fallback.")
                    else:
                        raise RuntimeError(f"Gemini API returned status {r.status_code}: {r.text}")
            except Exception as gemini_err:
                logger.warning(f"Gemini fallback also failed for Q{idx+1}: {gemini_err}. Trying OpenAI fallback...")
                try:
                    # Fall back to OpenAI
                    from openai import AsyncOpenAI
                    from app.core.config import settings
                    openai_client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
                    resp = await openai_client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role": "system", "content": "You are an expert exam solutions generator. Return only valid JSON."},
                            {"role": "user", "content": prompt}
                        ],
                        max_tokens=1024,
                        temperature=0.1
                    )
                    raw = resp.choices[0].message.content.strip()
                    import re as _re
                    clean = _re.sub(r'```json\s*', '', raw)
                    clean = _re.sub(r'\s*```', '', clean).strip()
                    first_brace = clean.find('{')
                    last_brace = clean.rfind('}')
                    if first_brace != -1 and last_brace != -1:
                        clean = clean[first_brace:last_brace+1]
                    res = _json.loads(clean)
                    success = True
                    logger.info(f"Successfully generated solution for Q{idx+1} using OpenAI fallback.")
                except Exception as oai_err:
                    logger.error(f"OpenAI fallback also failed for Q{idx+1}: {oai_err}")

        if success and isinstance(res, dict):
            if res.get("correct_answer"):
                q.correct_answer = res["correct_answer"]
            if res.get("explanation"):
                q.explanation = res["explanation"]
                total_explained += 1
            db.commit()

        # Brief delay to respect rate limits
        await asyncio.sleep(1.5)

    pyq_set.overview_generated = True
    db.commit()

    return {
        "success": True,
        "message": f"Overview generated for {total_explained} questions.",
        "total_explained": total_explained
    }


@router.post("/classroom/pyq-sets/{pyq_set_id}/generate-reel", tags=["Classroom PYQ"])
async def generate_pyq_reel(
    pyq_set_id: str,
    language: str = Form("English"),
    voice_id: Optional[str] = Form(None),
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    """
    Generate a max 5-minute Extension Reel summarising the entire PYQ set
    with correct answers and key explanations.
    """
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")

    questions = db.query(PYQQuestion).filter(
        PYQQuestion.pyq_set_id == pyq_set_id
    ).order_by(PYQQuestion.created_at.asc()).all()

    if not questions:
        raise HTTPException(400, "No questions in this PYQ set. Upload PDF first.")

    lang = language or "English"

    # Build a compact overview of up to 15 questions for the script
    q_summary = "\n".join(
        f"{i+1}. {q.question_text[:150]}{'...' if len(q.question_text)>150 else ''} → Ans: {q.correct_answer or 'See explanation'}"
        for i, q in enumerate(questions[:15])
    )

    from app.services.llm import generate_simple_response
    struct_prompt = f"""Create a structured scene-by-scene educational video reel script for a PYQ (Previous Year Questions) overview.

Exam Set: {pyq_set.name}
Language: {lang}

KEY QUESTIONS TO COVER:
{q_summary}

🎙️ Dialogue Rules:
1. If language is Hindi, write strictly in Devanagari Unicode. NEVER use Hinglish.
2. Spell out all numbers and acronyms in spoken words of the target language.

FORMAT FOR EACH SCENE:
🎬 Scene N (start-end sec)
🎙️ Dialogue: [Narration in {lang}, 20-30 words]
📸 Visuals / Footage: [Detailed English description]
🎥 Editing Notes: [Camera style]

Generate 15-20 scenes (max 5 minutes total = 300 seconds).
Do NOT add any markdown headers or extra text. Only output the scene blocks."""

    structured_script = await generate_simple_response(struct_prompt, "You are a professional PYQ video script writer. Output only the scene blocks.")

    from app.services.video_engine import assemble_advanced_reel
    res = await assemble_advanced_reel(
        structured_script,
        language=lang,
        voice_id=voice_id,
        bgm_style="cinematic"
    )

    video_url = res.get("video_url") if res else None
    if not video_url:
        raise HTTPException(500, "Reel generation failed")

    # Attempt R2 upload
    try:
        from app.services.r2_storage import upload_to_r2
        reel_token = secrets.token_hex(8)
        local_video_path = os.path.join(os.getcwd(), "uploads", "social", os.path.basename(video_url))
        r2_key = f"reels/pyq_{pyq_set_id}/reel_{reel_token}.mp4"
        r2_url = upload_to_r2(local_video_path, r2_key, "video/mp4")
        if r2_url:
            video_url = r2_url
    except Exception as r2_err:
        logger.error(f"PYQ reel R2 upload failed (using local): {r2_err}")

    return {
        "success": True,
        "video_url": video_url,
        "script": structured_script,
        "questions_covered": min(len(questions), 15)
    }


@router.get("/classroom/pyq-sets/{pyq_set_id}/reels", tags=["Classroom PYQ"])
async def list_pyq_reels(pyq_set_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """List all reels generated for a PYQ set."""
    from app.core.models import PYQSet, SocialContent
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ Set not found")
        
    # Query all reels for this client
    reels = db.query(SocialContent).filter(
        SocialContent.client_id == client["client_id"],
        SocialContent.content_type == "reel"
    ).order_by(SocialContent.created_at.desc()).all()
    
    # Filter by pyq_set_id in metadata
    filtered = []
    for r in reels:
        meta = r.metadata_info
        if meta.get("pyq_set_id") == pyq_set_id:
            filtered.append({
                "reel_id": r.content_id,
                "pyq_set_id": pyq_set_id,
                "media_url": r.media_url,
                "script": meta.get("script", r.body or ""),
                "created_at": r.created_at.isoformat() if r.created_at else None
            })
            
    return {"success": True, "reels": filtered}


@router.delete("/classroom/pyq-sets/reels/{reel_id}", tags=["Classroom PYQ"])
async def delete_pyq_reel(reel_id: str, client: dict = Depends(_require_client), db: Session = Depends(get_db)):
    """Delete a specific PYQ set reel."""
    from app.core.models import SocialContent
    reel = db.query(SocialContent).filter(
        SocialContent.content_id == reel_id,
        SocialContent.client_id == client["client_id"],
        SocialContent.content_type == "reel"
    ).first()
    if not reel:
        raise HTTPException(404, "Reel not found")
        
    db.delete(reel)
    db.commit()
    return {"success": True}


# ── Classroom Chatbot Endpoints ───────────────────────────────────────────────

class ChatRequest(BaseModel):
    question: str
    top_k: Optional[int] = 5


@router.post("/classroom/papers/{paper_id}/vectorize", tags=["Classroom Chatbot"])
async def vectorize_paper(
    paper_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    paper = db.query(PaperClassroom).join(Exam).filter(
        PaperClassroom.paper_id == paper_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")

    text_parts = []
    text_parts.append(f"Exam: {paper.exam.name}")
    text_parts.append(f"Exam Paper: {paper.name}")

    for subj in paper.subjects:
        text_parts.append(f"Subject: {subj.name}")
        for chap in subj.chapters:
            text_parts.append(f"Chapter: {chap.name} (Subject: {subj.name})")
            for topic in chap.topics:
                text_parts.append(f"Topic: {topic.name} (Chapter: {chap.name}, Subject: {subj.name})")
                if topic.description:
                    text_parts.append(f"Topic {topic.name} Description: {topic.description}")
                if topic.notes:
                    text_parts.append(f"Topic {topic.name} Notes:\n{topic.notes}")
                if topic.script:
                    text_parts.append(f"Topic {topic.name} Script:\n{topic.script}")
                for subtopic in topic.subtopics:
                    text_parts.append(f"Subtopic: {subtopic.name} (Topic: {topic.name}, Chapter: {chap.name})")
                    if subtopic.description:
                        text_parts.append(f"Subtopic {subtopic.name} Description: {subtopic.description}")
                    if subtopic.notes:
                        text_parts.append(f"Subtopic {subtopic.name} Notes:\n{subtopic.notes}")
                    if subtopic.script:
                        text_parts.append(f"Subtopic {subtopic.name} Script:\n{subtopic.script}")

    compiled_text = "\n\n".join(text_parts)
    
    if not compiled_text.strip():
        raise HTTPException(400, "No syllabus content found under this paper to vectorize.")

    from app.services.chunker import chunk_text
    from app.services.embedder import embed_texts
    from app.services.vector_store import get_vector_store

    chunks = chunk_text([(1, compiled_text)], source_file=f"Paper: {paper.name}")
    for c in chunks:
        c.paper_id = paper_id

    store = get_vector_store()
    store.delete_by_paper(paper_id)
    
    texts = [c.text for c in chunks]
    embeddings = embed_texts(texts)
    store.add_chunks(embeddings, chunks)

    return {"success": True, "total_chunks": len(chunks), "message": "Paper data successfully vectorized!"}


@router.post("/classroom/pyq-sets/{pyq_set_id}/vectorize", tags=["Classroom Chatbot"])
async def vectorize_pyq_set(
    pyq_set_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")

    questions = db.query(PYQQuestion).filter(
        PYQQuestion.pyq_set_id == pyq_set_id
    ).all()
    if not questions:
        raise HTTPException(400, "No questions found in this PYQ set to vectorize.")

    text_parts = [f"PYQ Set Name: {pyq_set.name}"]
    for idx, q in enumerate(questions):
        q_text = f"Question {idx+1}: {q.question_text}"
        if q.options:
            opts = ", ".join(q.options)
            q_text += f"\nOptions: {opts}"
        if q.correct_answer:
            q_text += f"\nCorrect Answer: {q.correct_answer}"
        if q.explanation:
            q_text += f"\nExplanation: {q.explanation}"
        text_parts.append(q_text)

    compiled_text = "\n\n---\n\n".join(text_parts)

    from app.services.chunker import chunk_text
    from app.services.embedder import embed_texts
    from app.services.vector_store import get_vector_store

    chunks = chunk_text([(1, compiled_text)], source_file=f"PYQ Set: {pyq_set.name}")
    for c in chunks:
        c.pyq_set_id = pyq_set_id

    store = get_vector_store()
    store.delete_by_pyq_set(pyq_set_id)
    
    texts = [c.text for c in chunks]
    embeddings = embed_texts(texts)
    store.add_chunks(embeddings, chunks)

    return {"success": True, "total_chunks": len(chunks), "message": "PYQ set successfully vectorized!"}


# ── Classroom Chatbot Endpoints ───────────────────────────────────────────────

class ChatRequest(BaseModel):
    question: str
    top_k: Optional[int] = 5


# Cooldown memory for rate-limited models to prevent repeated latency on subsequent requests
MODEL_COOLDOWNS = {}
COOLDOWN_DURATION = 300  # 5 minutes in seconds

async def _chat_llm_with_fallback(prompt: str, system_prompt: str, max_tokens: int = 1024) -> str:
    """
    Try primary Groq model first; auto-fallback on rate-limit / errors:
      1. llama-3.3-70b-versatile  (primary, 100K TPD)
      2. llama-3.1-8b-instant     (separate quota, fastest)
      3. gemini-2.5-flash         (generous free tier)
    Uses a cooldown dictionary to skip models that are currently rate-limited.
    """
    import time
    from app.services.llm import generate_simple_response
    from app.services.llm import set_runtime_provider, get_active_provider, get_active_model
    from app.core.config import settings

    GROQ_FALLBACK_MODELS = [
        "llama-3.3-70b-versatile",   # primary
        "llama-3.1-8b-instant",      # fallback 1 — separate TPD quota, very fast
    ]

    last_error = None
    primary_provider = get_active_provider()
    now = time.time()

    # Try Groq models in order, skipping ones on active cooldown
    for model in GROQ_FALLBACK_MODELS:
        if model in MODEL_COOLDOWNS and MODEL_COOLDOWNS[model] > now:
            logger.info(f"Skipping model '{model}' due to active rate-limit cooldown (expires in {int(MODEL_COOLDOWNS[model] - now)}s).")
            continue

        # If fallback model is llama-3.1-8b-instant, enforce strict token/prompt size limit
        active_prompt = prompt
        if model == "llama-3.1-8b-instant" and len(active_prompt) > 12000:
            logger.info(f"Prompt length ({len(active_prompt)} chars) is large for llama-3.1-8b-instant. Truncating context...")
            header = None
            if "--- Mapped Exam Paper Context ---\n" in active_prompt:
                header = "--- Mapped Exam Paper Context ---\n"
            elif "--- Mapped PYQ Set Context ---\n" in active_prompt:
                header = "--- Mapped PYQ Set Context ---\n"
                
            if header and "\n---\n\n--- Chat History ---" in active_prompt:
                parts = active_prompt.split(header, 1)
                subparts = parts[1].split("\n---\n\n--- Chat History ---", 1)
                context_text = subparts[0]
                # Truncate context to ~4500 chars (approx 1100 tokens) to be extremely safe
                truncated_context = context_text[:4500] + "\n... (context truncated to fit model token limit)"
                active_prompt = parts[0] + header + truncated_context + "\n---\n\n--- Chat History ---" + subparts[1]

        try:
            old_model = get_active_model()
            set_runtime_provider("groq", settings.GROQ_API_KEY, model)
            result = await generate_simple_response(active_prompt, system_prompt, max_tokens=max_tokens)
            # Restore to original on success
            set_runtime_provider(primary_provider, "", "")  # revert
            return result
        except Exception as e:
            last_error = e
            err_str = str(e).lower()
            # If rate-limited or quota exceeded, trigger cooldown
            if any(term in err_str for term in ["rate_limit", "429", "quota", "tokens per day", "limit reached"]):
                logger.warning(f"Groq model '{model}' rate-limited. Setting 5-minute cooldown.")
                MODEL_COOLDOWNS[model] = time.time() + COOLDOWN_DURATION
            else:
                logger.warning(f"Groq model '{model}' failed ({e}). Setting 1-minute cooldown.")
                MODEL_COOLDOWNS[model] = time.time() + 60
            continue

    # Final fallback: Gemini Flash (generous free quota)
    try:
        logger.warning("All Groq models failed or on cooldown, falling back to Gemini Flash...")
        set_runtime_provider("gemini", settings.GEMINI_API_KEY, "gemini-2.5-flash")
        result = await generate_simple_response(prompt, system_prompt, max_tokens=max_tokens)
        set_runtime_provider(primary_provider, "", "")  # revert
        return result
    except Exception as e:
        last_error = e
        logger.error(f"All LLM fallbacks failed. Last error: {e}")
        set_runtime_provider(primary_provider, "", "")  # revert
        raise RuntimeError(f"All LLM providers exhausted. Last error: {e}")


@router.post("/classroom/papers/{paper_id}/chat", tags=["Classroom Chatbot"])
async def chat_classroom_paper(
    paper_id: str,
    req: ChatRequest,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    paper = db.query(PaperClassroom).join(Exam).filter(
        PaperClassroom.paper_id == paper_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")

    question = req.question.strip()
    
    # ── Instant interception: greetings ──────────────────────────────────────
    q_clean = question.lower().strip().rstrip("?").rstrip(".").rstrip("!")
    if q_clean in ["hi", "hii", "hello", "hey", "helo", "hllo"]:
        answer = "hii how are you"
        db.add(PaperChat(paper_id=paper_id, role="user", content=question, sources_json="[]"))
        db.add(PaperChat(paper_id=paper_id, role="assistant", content=answer, sources_json="[]"))
        db.commit()
        return {"success": True, "answer": answer, "sources": [], "context_found": False}

    # ── Instant interception: identity questions ──────────────────────────────
    _identity_triggers = [
        "who are you", "what is your name", "your name", "what are you",
        "aap kaun ho", "tumhara naam kya hai", "apna naam batao", "kaun ho tum",
        "introduce yourself", "tell me about yourself", "tum kaun ho",
        "aapka naam", "naam kya hai", "aap kya ho",
    ]
    if any(t in q_clean for t in _identity_triggers):
        exam_name = paper.exam.name if paper.exam else "Exam"
        paper_name = paper.name
        answer = f"I am {paper_name} of {exam_name} helper. I can answer questions related to the {paper_name} syllabus and content."
        db.add(PaperChat(paper_id=paper_id, role="user", content=question, sources_json="[]"))
        db.add(PaperChat(paper_id=paper_id, role="assistant", content=answer, sources_json="[]"))
        db.commit()
        return {"success": True, "answer": answer, "sources": [], "context_found": False}

    # Fetch last 6 messages of history
    history_msgs = db.query(PaperChat).filter(
        PaperChat.paper_id == paper_id
    ).order_by(PaperChat.timestamp.desc()).limit(6).all()
    history_msgs.reverse()

    history_str = ""
    for h in history_msgs:
        role_name = "User" if h.role == "user" else "Assistant"
        history_str += f"{role_name}: {h.content}\n"

    from app.services.embedder import embed_query
    from app.services.vector_store import get_vector_store
    from app.services.llm import build_context_and_sources, generate_simple_response
    import json

    search_query = question
    if any(ord(char) > 127 for char in question):
        try:
            translation_prompt = f"Translate the following Hindi query into a concise English search query for information retrieval. Only return the direct English translation, nothing else.\nQuery: {question}"
            translated = await _chat_llm_with_fallback(translation_prompt, "You are a translator. Only return the direct translation.", max_tokens=128)
            search_query = translated.strip()
            logger.info(f"Translated paper query: '{question}' -> '{search_query}'")
        except Exception as e:
            logger.error(f"Failed to translate paper query: {e}")

    query_emb = embed_query(search_query)
    store = get_vector_store()
    results = store.search_by_paper(query_emb, paper_id, top_k=req.top_k)
    
    valid_results = [(chunk, score) for chunk, score in results if score >= 0.25]
    
    context_found = bool(valid_results)
    prefix_msg = ""
    
    if not context_found:
        prefix_msg = "ye apke data se nhi h apne traf se bata raha hu\n\n"
        context = "No relevant context was found in the indexed documents. Please answer the user's question using your general knowledge."
        sources_data = []
    else:
        context, sources_data = build_context_and_sources(valid_results)

    system_prompt = (
        f"You are a helpful educational AI assistant helper for the Exam Paper: '{paper.name}' (Exam: '{paper.exam.name}').\n"
        f"Answer the user's question as accurately as possible based on the provided context.\n"
        f"Answer strictly in the SAME language that the user used to ask their question.\n"
        f"Cite sources if retrieved from context.\n"
    )

    prompt = (
        f"--- Mapped Exam Paper Context ---\n{context}\n---\n\n"
        f"--- Chat History ---\n{history_str}\n---\n\n"
        f"Question: {question}"
    )

    try:
        raw_answer = await _chat_llm_with_fallback(prompt, system_prompt, max_tokens=1024)
        answer = prefix_msg + raw_answer
    except Exception as e:
        logger.error(f"Paper chat LLM failed (all fallbacks): {e}")
        raise HTTPException(502, f"LLM error: {e}")

    srcs_json = json.dumps([s.model_dump() for s in sources_data], default=str)
    db.add(PaperChat(paper_id=paper_id, role="user", content=question, sources_json="[]"))
    db.add(PaperChat(paper_id=paper_id, role="assistant", content=answer, sources_json=srcs_json))
    db.commit()

    return {
        "success": True,
        "question": question,
        "answer": answer,
        "sources": [s.model_dump() for s in sources_data],
        "context_found": context_found
    }


@router.get("/classroom/papers/{paper_id}/chat/history", tags=["Classroom Chatbot"])
async def get_paper_chat_history(
    paper_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    paper = db.query(PaperClassroom).join(Exam).filter(
        PaperClassroom.paper_id == paper_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")

    msgs = db.query(PaperChat).filter(
        PaperChat.paper_id == paper_id
    ).order_by(PaperChat.timestamp.asc()).all()

    from app.services.vector_store import get_vector_store
    store = get_vector_store()
    is_vectorized = any(getattr(m, "paper_id", None) == paper_id for m in store.metadata)

    return {"success": True, "history": [m.to_dict() for m in msgs], "is_vectorized": is_vectorized}


@router.delete("/classroom/papers/{paper_id}/chat/history", tags=["Classroom Chatbot"])
async def clear_paper_chat_history(
    paper_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    paper = db.query(PaperClassroom).join(Exam).filter(
        PaperClassroom.paper_id == paper_id,
        Exam.client_id == client["client_id"]
    ).first()
    if not paper:
        raise HTTPException(404, "Paper not found or access denied")

    db.query(PaperChat).filter(PaperChat.paper_id == paper_id).delete(synchronize_session=False)
    db.commit()

    return {"success": True, "message": "Chat history cleared!"}


@router.post("/classroom/pyq-sets/{pyq_set_id}/chat", tags=["Classroom Chatbot"])
async def chat_classroom_pyq_set(
    pyq_set_id: str,
    req: ChatRequest,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ set not found")

    question = req.question.strip()
    
    # ── Instant interception: greetings ──────────────────────────────────────
    q_clean = question.lower().strip().rstrip("?").rstrip(".").rstrip("!")
    if q_clean in ["hi", "hii", "hello", "hey", "helo", "hllo"]:
        answer = "hii how are you"
        db.add(PYQChat(pyq_set_id=pyq_set_id, role="user", content=question, sources_json="[]"))
        db.add(PYQChat(pyq_set_id=pyq_set_id, role="assistant", content=answer, sources_json="[]"))
        db.commit()
        return {"success": True, "answer": answer, "sources": [], "context_found": False}

    # ── Instant interception: identity questions ──────────────────────────────
    _identity_triggers = [
        "who are you", "what is your name", "your name", "what are you",
        "aap kaun ho", "tumhara naam kya hai", "apna naam batao", "kaun ho tum",
        "introduce yourself", "tell me about yourself", "tum kaun ho",
        "aapka naam", "naam kya hai", "aap kya ho",
    ]
    if any(t in q_clean for t in _identity_triggers):
        answer = f"I am {pyq_set.name} helper. I can answer questions about the questions and solutions in this PYQ set."
        db.add(PYQChat(pyq_set_id=pyq_set_id, role="user", content=question, sources_json="[]"))
        db.add(PYQChat(pyq_set_id=pyq_set_id, role="assistant", content=answer, sources_json="[]"))
        db.commit()
        return {"success": True, "answer": answer, "sources": [], "context_found": False}

    # Fetch last 6 messages of history
    history_msgs = db.query(PYQChat).filter(
        PYQChat.pyq_set_id == pyq_set_id
    ).order_by(PYQChat.timestamp.desc()).limit(6).all()
    history_msgs.reverse()

    history_str = ""
    for h in history_msgs:
        role_name = "User" if h.role == "user" else "Assistant"
        history_str += f"{role_name}: {h.content}\n"

    from app.services.embedder import embed_query
    from app.services.vector_store import get_vector_store
    from app.services.llm import build_context_and_sources, generate_simple_response
    import json

    search_query = question
    if any(ord(char) > 127 for char in question):
        try:
            translation_prompt = f"Translate the following Hindi query into a concise English search query for information retrieval. Only return the direct English translation, nothing else.\nQuery: {question}"
            translated = await _chat_llm_with_fallback(translation_prompt, "You are a translator. Only return the direct translation.", max_tokens=128)
            search_query = translated.strip()
            logger.info(f"Translated PYQ query: '{question}' -> '{search_query}'")
        except Exception as e:
            logger.error(f"Failed to translate PYQ query: {e}")

    query_emb = embed_query(search_query)
    store = get_vector_store()
    results = store.search_by_pyq_set(query_emb, pyq_set_id, top_k=req.top_k)
    
    valid_results = [(chunk, score) for chunk, score in results if score >= 0.25]
    
    context_found = bool(valid_results)
    prefix_msg = ""
    
    if not context_found:
        prefix_msg = "ye apke data se nhi h apne traf se bata raha hu\n\n"
        context = "No relevant context was found in the indexed documents. Please answer the user's question using your general knowledge."
        sources_data = []
    else:
        context, sources_data = build_context_and_sources(valid_results)

    system_prompt = (
        f"You are a helpful educational AI assistant helper for the PYQ Set: '{pyq_set.name}'.\n"
        f"Answer the user's question as accurately as possible based on the provided context.\n"
        f"Answer strictly in the SAME language that the user used to ask their question.\n"
        f"Cite sources if retrieved from context.\n"
    )

    prompt = (
        f"--- Mapped PYQ Set Context ---\n{context}\n---\n\n"
        f"--- Chat History ---\n{history_str}\n---\n\n"
        f"Question: {question}"
    )

    try:
        raw_answer = await _chat_llm_with_fallback(prompt, system_prompt, max_tokens=1024)
        answer = prefix_msg + raw_answer
    except Exception as e:
        logger.error(f"PYQ chat LLM failed (all fallbacks): {e}")
        raise HTTPException(502, f"LLM error: {e}")

    srcs_json = json.dumps([s.model_dump() for s in sources_data], default=str)
    db.add(PYQChat(pyq_set_id=pyq_set_id, role="user", content=question, sources_json="[]"))
    db.add(PYQChat(pyq_set_id=pyq_set_id, role="assistant", content=answer, sources_json=srcs_json))
    db.commit()

    return {
        "success": True,
        "question": question,
        "answer": answer,
        "sources": [s.model_dump() for s in sources_data],
        "context_found": context_found
    }


@router.get("/classroom/pyq-sets/{pyq_set_id}/chat/history", tags=["Classroom Chatbot"])
async def get_pyq_chat_history(
    pyq_set_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ Set not found")

    msgs = db.query(PYQChat).filter(
        PYQChat.pyq_set_id == pyq_set_id
    ).order_by(PYQChat.timestamp.asc()).all()

    from app.services.vector_store import get_vector_store
    store = get_vector_store()
    is_vectorized = any(getattr(m, "pyq_set_id", None) == pyq_set_id for m in store.metadata)

    return {"success": True, "history": [m.to_dict() for m in msgs], "is_vectorized": is_vectorized}


@router.delete("/classroom/pyq-sets/{pyq_set_id}/chat/history", tags=["Classroom Chatbot"])
async def clear_pyq_chat_history(
    pyq_set_id: str,
    client: dict = Depends(_require_client),
    db: Session = Depends(get_db)
):
    pyq_set = db.query(PYQSet).filter(
        PYQSet.pyq_set_id == pyq_set_id,
        PYQSet.client_id == client["client_id"]
    ).first()
    if not pyq_set:
        raise HTTPException(404, "PYQ Set not found")

    db.query(PYQChat).filter(PYQChat.pyq_set_id == pyq_set_id).delete(synchronize_session=False)
    db.commit()

    return {"success": True, "message": "Chat history cleared!"}


# ── ElevenLabs TTS Proxy ──────────────────────────────────────────────────────

class TTSRequest(BaseModel):
    text: str
    voice_id: Optional[str] = "ypnkIsDASPgHZuanrF0q"  # Rudra (Hindi male)


@router.post("/classroom/tts/speak", tags=["Classroom Chatbot"])
async def elevenlabs_tts_proxy(
    req: TTSRequest,
    client: dict = Depends(_require_client),
):
    """
    Proxy text-to-speech via ElevenLabs Rudra voice.
    Returns audio/mpeg binary so frontend never exposes the API key.
    """
    import httpx
    from fastapi.responses import Response
    from app.core.config import settings

    api_key = settings.ELEVENLABS_API_KEY
    if not api_key:
        raise HTTPException(500, "ElevenLabs API key not configured on server.")

    voice_id = req.voice_id or "ypnkIsDASPgHZuanrF0q"
    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"

    # Trim to ElevenLabs character limit (2500 chars max for safety)
    text = req.text.strip()
    # Trim for voice — flash model is fastest with shorter text
    text = req.text.strip()
    if len(text) > 250:
        # Cut at word boundary
        cut = text.rfind(' ', 0, 250)
        text = text[:cut if cut > 150 else 250]

    # eleven_flash_v2_5 = ElevenLabs fastest model (~0.3s latency, supports Hindi/English)
    payload = {
        "text": text,
        "model_id": "eleven_flash_v2_5",
        "voice_settings": {
            "stability": 0.45,
            "similarity_boost": 0.80,
            "style": 0.0,
            "use_speaker_boost": True
        }
    }
    headers = {
        "xi-api-key": api_key,
        "Content-Type": "application/json",
        "Accept": "audio/mpeg"
    }

    # Use smaller MP3 (22050Hz 32kbps) for faster download → add as query param
    tts_url = url + "?output_format=mp3_22050_32"

    try:
        async with httpx.AsyncClient(timeout=15.0) as hc:
            r = await hc.post(tts_url, json=payload, headers=headers)
            if r.status_code == 200:
                return Response(content=r.content, media_type="audio/mpeg")
            else:
                logger.error(f"ElevenLabs TTS error {r.status_code}: {r.text}")
                raise HTTPException(502, f"ElevenLabs error: {r.status_code}")
    except httpx.RequestError as e:
        logger.error(f"ElevenLabs TTS request failed: {e}")
        raise HTTPException(502, f"TTS request failed: {e}")


# Shared HTTP client for connection pooling to ElevenLabs (skips TCP/SSL handshake latency)
_tts_http_client = None

def get_tts_http_client():
    global _tts_http_client
    if _tts_http_client is None:
        import httpx
        _tts_http_client = httpx.AsyncClient(timeout=30.0)
    return _tts_http_client


@router.get("/classroom/tts/speak", tags=["Classroom Chatbot"])
async def elevenlabs_tts_proxy_get(
    text: str,
    voice_id: Optional[str] = "ypnkIsDASPgHZuanrF0q",
    client: dict = Depends(_require_client),
):
    """
    Stream text-to-speech via ElevenLabs voice.
    Returns audio/mpeg binary stream for ultra-low latency playback.
    """
    from fastapi.responses import StreamingResponse
    from app.core.config import settings

    api_key = settings.ELEVENLABS_API_KEY
    if not api_key:
        raise HTTPException(500, "ElevenLabs API key not configured on server.")

    # Trim to ElevenLabs character limit (250 chars max for speed)
    text = text.strip()
    if len(text) > 250:
        cut = text.rfind(' ', 0, 250)
        text = text[:cut if cut > 150 else 250]

    payload = {
        "text": text,
        "model_id": "eleven_flash_v2_5",
        "voice_settings": {
            "stability": 0.45,
            "similarity_boost": 0.80,
            "style": 0.0,
            "use_speaker_boost": False
        }
    }
    headers = {
        "xi-api-key": api_key,
        "Content-Type": "application/json",
        "Accept": "audio/mpeg"
    }

    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"
    # Use output_format and optimize_streaming_latency=4 for max speed
    tts_url = url + "?output_format=mp3_22050_32&optimize_streaming_latency=4"

    async def audio_generator():
        try:
            hc = get_tts_http_client()
            async with hc.stream("POST", tts_url, json=payload, headers=headers) as r:
                if r.status_code == 200:
                    async for chunk in r.aiter_bytes(chunk_size=1024):
                        yield chunk
                else:
                    err_text = await r.aread()
                    logger.error(f"ElevenLabs TTS error {r.status_code}: {err_text.decode('utf-8', errors='ignore')}")
        except Exception as e:
            logger.error(f"ElevenLabs streaming failed: {e}")

    return StreamingResponse(audio_generator(), media_type="audio/mpeg")

